import os
import sys

# Ensure app directory and project root are always in sys.path for direct imports (e.g. gunicorn app.app:app)
_app_dir = os.path.dirname(os.path.abspath(__file__))
_root_dir = os.path.dirname(_app_dir)
for _p in [_app_dir, _root_dir]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

import csv
import io
import json
import re
import secrets
import subprocess
import threading
import time
import uuid
import zipfile
from collections import OrderedDict
from datetime import datetime, timedelta

import pymysql
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, send_from_directory, session, stream_with_context
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

import file_scraper_runner
import files_repo
import reports_repo
import job_manager
from auth import (
    VALID_ROLES,
    bit_to_bool,
    get_user_by_email,
    get_user_by_id,
    has_superadmin,
    hash_password,
    list_active_users,
    list_deleted_users,
    login_required_api,
    login_required_page,
    require_csrf,
    role_required_api,
    role_required_page,
    serialize_user,
    verify_password,
)
from db import get_connection
from mailer import send_email
from scraper_status_utils import build_status_summary, get_xlsx_info, parse_status_line

EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

# This file lives in app/, but templates/, static/, scrapers/, and the scraper's
# xlsx output all live at the project root (one level up) -- everything below
# that needs a filesystem path is anchored to BASE_DIR, not this file's own directory.
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# scraper_config.py and scraper_input.py live in scrapers/, not app/ -- put that
# folder on sys.path so they can be bare-imported the same way auth/db/mailer
# are (see the module docstring note about sys.path[0] in CLAUDE.md).
sys.path.insert(0, os.path.join(BASE_DIR, 'scrapers'))
from scraper_config import SCRIPT_MAP  # noqa: E402
from scraper_input import (  # noqa: E402
    InputError,
    build_entries,
    extract_input_source,
    format_invalid_url_message,
    format_unsupported_message,
    parse_csv_urls,
    parse_text_urls,
    validate_url_list,
)

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static'),
)

# High-Performance HTTP Response Compression (Brotli + Gzip)
try:
    from flask_compress import Compress
    Compress(app)
    app.config['COMPRESS_ALGORITHM'] = ['brotli', 'gzip', 'deflate']
    app.config['COMPRESS_MIN_SIZE'] = 500
except ImportError:
    pass

# A random fallback key here would change on every process restart (e.g. Render's
# free-tier spin-down/cold-start), invalidating every existing session cookie and
# making the app look like it "reset" on refresh. Set FLASK_SECRET_KEY in the
# hosting environment so sessions survive restarts.
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-only-insecure-key-set-FLASK_SECRET_KEY-in-production')
app.permanent_session_lifetime = timedelta(days=7)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 604800  # 7 days browser cache for static files


@app.after_request
def add_performance_headers(response):
    """Adds caching headers for static assets and enables keep-alive."""
    if request.path.startswith('/static/'):
        # Cache static CSS, JS, images for 7 days with revalidation
        response.headers['Cache-Control'] = 'public, max-age=604800, stale-while-revalidate=86400'
    return response


TMP_SCRAPERS_DIR = os.path.join(BASE_DIR, 'tmp', 'scrapers')


class ScraperSession:
    """Per-browser-session scraper state, so concurrent users don't see or control each other's runs.

    job_status is a one-shot state machine:
      idle -> running -> completed_unseen/failed_unseen -> idle (archived)
    The completed/failed_unseen states are reported exactly once (by /scraper-status)
    then immediately archived back to idle, so a job that finished is never treated
    as "active" again on a later poll or a fresh page load (browser refresh).

    A single job can involve up to three different scraper scripts (one per
    detected URL type). pending_groups holds one entry per type still to run;
    they execute sequentially in a background thread (_run_job_groups), with
    `process` always pointing at whichever group's subprocess is currently
    active so /stop-scraper and /scraper-status keep working unchanged.
    """

    def __init__(self):
        self.lock = threading.Lock()
        self.process = None
        self.thread = None
        self.url_statuses = []
        self.job_id = None
        self.job_status = 'idle'
        self.stopped = False
        self.output_file = None
        self.pending_groups = []
        self.skipped = {'invalid': [], 'unsupported': []}


scraper_sessions = {}
scraper_sessions_lock = threading.Lock()


def get_session_id():
    if 'sid' not in session:
        session.permanent = True
        session['sid'] = secrets.token_hex(16)
    return session['sid']


def get_scraper_session():
    session_id = get_session_id()
    with scraper_sessions_lock:
        state = scraper_sessions.get(session_id)
        if state is None:
            state = ScraperSession()
            scraper_sessions[session_id] = state
    return state


# @app.route('/')
# def index():
#     return render_template("Dashboard.html", page="Dashboard")


@app.route('/scraperpage')
@login_required_page
def Scrap():
    file_id = request.args.get('fileId')
    if not file_id:
        user_id = session.get('user_id')
        active_map = job_manager.get_all_active_jobs_map()
        running_for_user = [fid for fid, info in active_map.items() if info.get('user_id') == user_id]
        if running_for_user:
            return redirect(f'/scraperpage?fileId={running_for_user[0]}')
        if active_map:
            return redirect(f'/scraperpage?fileId={list(active_map.keys())[0]}')

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT file_id FROM logTbl ORDER BY id DESC LIMIT 1")
                latest_log = cursor.fetchone()
                if latest_log and latest_log.get('file_id'):
                    return redirect(f'/scraperpage?fileId={latest_log["file_id"]}')
                cursor.execute("SELECT file_id FROM fileTbl WHERE is_deleted = 0 ORDER BY file_id ASC LIMIT 1")
                first_file = cursor.fetchone()
                if first_file:
                    return redirect(f'/scraperpage?fileId={first_file["file_id"]}')
        finally:
            conn.close()

    return render_template("Scrap.html", page="scraping")


@app.route('/')
@login_required_page
def files_page():
    return render_template('files.html', page='files')


@app.route('/docs/scraper')
@login_required_page
def scraper_guide_page():
    """Documentation only -- explains the existing scraper contract
    (argv output/input paths, URL_STATUS protocol, FEEDS/xlsx export) using
    scrapers/pitstoparabiabycsv.py as the reference implementation. Does not
    change the scraper system, database, or upload flow in any way.
    """
    return render_template('scraper_guide.html', page='docs')


@app.route('/login', methods=['GET'])
def login_page():
    if 'user_id' in session:
        return redirect('/')
    return render_template('login.html')


LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCKOUT_SECONDS = 15 * 60

_login_attempts_lock = threading.Lock()
_login_attempts = {}  # email -> {'count': int, 'locked_until': monotonic time or None}


def _login_lockout_remaining(email):
    """Returns seconds left in an active lockout for this email, or None if it can try again."""
    now = time.monotonic()
    with _login_attempts_lock:
        entry = _login_attempts.get(email)
        if not entry or not entry['locked_until']:
            return None
        remaining = entry['locked_until'] - now
        if remaining <= 0:
            # Lockout has expired -- give this email a fresh set of attempts.
            _login_attempts.pop(email, None)
            return None
        return remaining


def _record_login_failure(email):
    with _login_attempts_lock:
        entry = _login_attempts.setdefault(email, {'count': 0, 'locked_until': None})
        entry['count'] += 1
        if entry['count'] >= LOGIN_MAX_ATTEMPTS:
            entry['count'] = 0
            entry['locked_until'] = time.monotonic() + LOGIN_LOCKOUT_SECONDS


def _clear_login_failures(email):
    with _login_attempts_lock:
        _login_attempts.pop(email, None)


@app.route('/login', methods=['POST'])
def login_submit():
    data = request.get_json(silent=True) or request.form
    email = (data.get('email') or '').strip()
    password = data.get('password') or ''
    remember = bool(data.get('remember'))

    if not email or not password:
        return jsonify({'error': 'Email and password are required.'}), 400

    remaining = _login_lockout_remaining(email)
    if remaining is not None:
        minutes = max(1, round(remaining / 60))
        return jsonify({'error': f'Too many failed attempts. Try again in {minutes} minute(s).'}), 429

    user = get_user_by_email(email)
    if not user or bit_to_bool(user['IsDeleted']) or not verify_password(password, user['password']):
        _record_login_failure(email)
        return jsonify({'error': 'Invalid email or password.'}), 401

    if not bit_to_bool(user['Status']):
        return jsonify({'error': 'This account has been disabled. Contact an administrator.'}), 403

    _clear_login_failures(email)

    # Regenerate the whole session (fresh sid + auth claims) on every successful
    # login so a pre-login session id can never be reused post-login (session fixation).
    session.clear()
    session.permanent = remember
    session['sid'] = secrets.token_hex(16)
    session['user_id'] = user['userid']
    session['name'] = user['Name']
    session['email'] = user['Email']
    session['role'] = user['Role']
    session['csrf_token'] = secrets.token_hex(16)

    return jsonify({'redirect': '/'})


@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'redirect': '/login'})


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password_page():
    return redirect('/login')


@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password_page():
    return redirect('/login')


@app.route('/api/me')
@login_required_api
def api_me():
    user = get_user_by_id(session['user_id'])
    if not user:
        session.clear()
        return jsonify({'error': 'Authentication required.'}), 401
    return jsonify({'user': serialize_user(user), 'csrfToken': session.get('csrf_token')})


@app.route('/api/profile', methods=['PUT'])
@login_required_api
@require_csrf
def api_update_profile():
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    avatar = data.get('avatar')
    avatar = avatar.strip() if isinstance(avatar, str) else None
    avatar = avatar or None

    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    if not email or not EMAIL_RE.match(email):
        return jsonify({'error': 'A valid email is required.'}), 400
    if avatar and len(avatar) > 500:
        return jsonify({'error': 'Avatar URL is too long.'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            try:
                cursor.execute(
                    'UPDATE userTbl SET Name = %s, Email = %s, avatar = %s WHERE userid = %s',
                    (name, email, avatar, session['user_id']),
                )
            except pymysql.err.IntegrityError:
                return jsonify({'error': 'That email is already in use.'}), 409
    finally:
        conn.close()

    session['name'] = name
    session['email'] = email

    user = get_user_by_id(session['user_id'])
    return jsonify({'user': serialize_user(user)})


@app.route('/api/profile/avatar', methods=['DELETE'])
@login_required_api
@require_csrf
def api_remove_avatar():
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute('UPDATE userTbl SET avatar = NULL WHERE userid = %s', (session['user_id'],))
    finally:
        conn.close()

    user = get_user_by_id(session['user_id'])
    return jsonify({'user': serialize_user(user)})


@app.route('/api/change-password', methods=['POST'])
@login_required_api
@require_csrf
def api_change_password():
    fail_count = session.get('pwd_fail_count', 0)
    if fail_count >= 5:
        session.clear()
        return jsonify({'error': 'Too many failed attempts. Please log in again.'}), 429

    data = request.get_json(silent=True) or {}
    current_password = data.get('current_password') or ''
    new_password = data.get('new_password') or ''
    confirm_password = data.get('confirm_password') or ''

    if not current_password or not new_password or not confirm_password:
        return jsonify({'error': 'All fields are required.'}), 400
    if new_password != confirm_password:
        return jsonify({'error': 'New password and confirmation do not match.'}), 400
    if len(new_password) < 8:
        return jsonify({'error': 'New password must be at least 8 characters.'}), 400

    user = get_user_by_id(session['user_id'])
    if not user or not verify_password(current_password, user['password']):
        session['pwd_fail_count'] = fail_count + 1
        return jsonify({'error': 'Current password is incorrect.'}), 400

    if verify_password(new_password, user['password']):
        return jsonify({'error': 'New password must be different from the current password.'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                'UPDATE userTbl SET password = %s WHERE userid = %s',
                (hash_password(new_password), session['user_id']),
            )
    finally:
        conn.close()

    session.pop('pwd_fail_count', None)
    # Rotate identifiers after a credential change so a stolen pre-change
    # session/CSRF token pairing can't be replayed.
    session['sid'] = secrets.token_hex(16)
    session['csrf_token'] = secrets.token_hex(16)

    return jsonify({'message': 'Password changed successfully.'})


@app.route('/Admin')
@login_required_page
@role_required_page('SuperAdmin', 'Admin')
def admin_page():
    return render_template('admin.html', page='admin')


@app.route('/trash')
@login_required_page
@role_required_page('SuperAdmin', 'Admin')
def trash_page():
    return render_template('trash.html', page='trash')


@app.route('/reports')
@login_required_page
@role_required_page('SuperAdmin')
def reports_page():
    return render_template('reports.html', page='reports')


@app.route('/api/admin/users', methods=['GET'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
def api_admin_list_users():
    # Soft-deleted accounts never appear in the main User Management list --
    # they only show up on the Trash page (see api_admin_list_trash).
    return jsonify({'users': [serialize_user(u) for u in list_active_users()]})


@app.route('/api/admin/users/trash', methods=['GET'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
def api_admin_list_trash():
    return jsonify({'users': [serialize_user(u) for u in list_deleted_users()]})


@app.route('/api/admin/users', methods=['POST'])
@login_required_api
@role_required_api('SuperAdmin')
@require_csrf
def api_admin_create_user():
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    password = data.get('password') or ''
    role = (data.get('role') or '').strip()
    status = data.get('status', True)

    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    if not email or not EMAIL_RE.match(email):
        return jsonify({'error': 'A valid email is required.'}), 400
    if not password or len(password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters.'}), 400
    if role not in VALID_ROLES:
        return jsonify({'error': f"Role must be one of: {', '.join(VALID_ROLES)}."}), 400
    # Exactly one SuperAdmin may ever exist in the system.
    if role == 'SuperAdmin' and has_superadmin():
        return jsonify({'error': 'A SuperAdmin already exists. Only one SuperAdmin account is allowed.'}), 409

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            try:
                cursor.execute(
                    'INSERT INTO userTbl (Name, Email, password, Status, IsDeleted, Role) '
                    'VALUES (%s, %s, %s, %s, 0, %s)',
                    (name, email, hash_password(password), 1 if status else 0, role),
                )
            except pymysql.err.IntegrityError:
                return jsonify({'error': 'A user with that email already exists.'}), 409
            new_user_id = cursor.lastrowid
    finally:
        conn.close()

    return jsonify({'user': serialize_user(get_user_by_id(new_user_id))}), 201


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
@require_csrf
def api_admin_update_user(user_id):
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found.'}), 404

    actor_role = session.get('role')
    # An Admin can manage other Admins/Users, but never touch a SuperAdmin's
    # account -- that's reserved for SuperAdmins only.
    if target['Role'] == 'SuperAdmin' and actor_role != 'SuperAdmin':
        return jsonify({'error': 'Only a SuperAdmin can modify a SuperAdmin account.'}), 403

    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    role = (data.get('role') or '').strip()
    password = data.get('password') or ''
    status = data.get('status', True)

    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    if not email or not EMAIL_RE.match(email):
        return jsonify({'error': 'A valid email is required.'}), 400
    if role not in VALID_ROLES:
        return jsonify({'error': f"Role must be one of: {', '.join(VALID_ROLES)}."}), 400
    if role == 'SuperAdmin' and actor_role != 'SuperAdmin':
        return jsonify({'error': 'Only a SuperAdmin can grant the SuperAdmin role.'}), 403
    # Exactly one SuperAdmin may ever exist -- only exempt this check when the
    # target already holds the role (i.e. this edit isn't creating a new one).
    if role == 'SuperAdmin' and target['Role'] != 'SuperAdmin' and has_superadmin():
        return jsonify({'error': 'A SuperAdmin already exists. Only one SuperAdmin account is allowed.'}), 409
    if password and len(password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters.'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            try:
                if password:
                    cursor.execute(
                        'UPDATE userTbl SET Name = %s, Email = %s, Role = %s, Status = %s, password = %s '
                        'WHERE userid = %s',
                        (name, email, role, 1 if status else 0, hash_password(password), user_id),
                    )
                else:
                    cursor.execute(
                        'UPDATE userTbl SET Name = %s, Email = %s, Role = %s, Status = %s WHERE userid = %s',
                        (name, email, role, 1 if status else 0, user_id),
                    )
            except pymysql.err.IntegrityError:
                return jsonify({'error': 'That email is already in use.'}), 409
    finally:
        conn.close()

    updated = get_user_by_id(user_id)
    # Keep the acting user's own session in sync if they just edited themselves.
    if user_id == session.get('user_id'):
        session['name'] = updated['Name']
        session['email'] = updated['Email']
        session['role'] = updated['Role']

    return jsonify({'user': serialize_user(updated)})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
@require_csrf
def api_admin_delete_user(user_id):
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found.'}), 404

    # A SuperAdmin account can never be deleted -- not by another SuperAdmin,
    # and not by an Admin.
    if target['Role'] == 'SuperAdmin':
        return jsonify({'error': 'SuperAdmin accounts can never be deleted.'}), 403

    # No one -- SuperAdmin or Admin -- can delete their own account.
    if user_id == session.get('user_id'):
        return jsonify({'error': 'You cannot delete your own account.'}), 403

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Soft delete (matches how login/forgot-password already treat
            # IsDeleted) rather than removing the row outright -- this is what
            # moves the user into Trash.
            cursor.execute(
                'UPDATE userTbl SET IsDeleted = 1, deleted_at = UTC_TIMESTAMP() WHERE userid = %s',
                (user_id,),
            )
    finally:
        conn.close()

    return jsonify({'message': 'User deleted.'})


@app.route('/api/admin/users/<int:user_id>/recover', methods=['POST'])
@login_required_api
@role_required_api('SuperAdmin')
@require_csrf
def api_admin_recover_user(user_id):
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found.'}), 404
    if not bit_to_bool(target['IsDeleted']):
        return jsonify({'error': 'This account is not deleted.'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute('UPDATE userTbl SET IsDeleted = 0, deleted_at = NULL WHERE userid = %s', (user_id,))
    finally:
        conn.close()

    return jsonify({'message': 'User recovered.'})


def _record_status_line(state, cleaned_line):
    parsed_status = parse_status_line(cleaned_line)
    if not parsed_status:
        return
    with state.lock:
        existing = next((item for item in state.url_statuses if item['url'] == parsed_status['url']), None)
        if existing:
            existing['status'] = parsed_status['status']
            if parsed_status.get('parent'):
                existing['parent'] = parsed_status['parent']
            if parsed_status.get('type'):
                existing['type'] = parsed_status['type']
        else:
            state.url_statuses.append({
                'url': parsed_status['url'],
                'status': parsed_status['status'],
                'parent': parsed_status.get('parent') or '',
                'type': parsed_status.get('type') or 'root',
            })


def _merge_xlsx_outputs(source_paths, destination):
    """Concatenates the data rows of one-or-more group output workbooks (each
    produced by a different scraper script but sharing the same column schema)
    into a single workbook at `destination`, so the dashboard's one download
    link keeps working regardless of how many scrapers actually ran.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    header_written = False

    for path in source_paths:
        try:
            src_wb = load_workbook(path, read_only=True)
        except Exception:
            continue
        try:
            rows = src_wb.active.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            if not header_written:
                ws.append(list(header))
                header_written = True
            for row in rows:
                ws.append(list(row))
        finally:
            src_wb.close()

    if not header_written:
        ws.append(['No data'])

    os.makedirs(os.path.dirname(destination), exist_ok=True)
    wb.save(destination)


def _run_job_groups(state):
    """Runs each detected-type group's scraper subprocess in turn (brand, then
    sitemap, then listing/product -- whatever the job actually contains),
    merging their stdout into the same url_statuses list the frontend already
    knows how to render as one tree, then merges their xlsx outputs into the
    job's single downloadable file.
    """
    group_outputs = []
    had_failure = False

    for group in state.pending_groups:
        with state.lock:
            if state.stopped:
                break

        process = subprocess.Popen(
            [sys.executable, '-u', group['script_path'], group['output_path'], group['input_path']],
            cwd=BASE_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        with state.lock:
            state.process = process

        for line in iter(process.stdout.readline, ''):
            if not line:
                break
            _record_status_line(state, line.rstrip('\n'))
        process.stdout.close()
        process.wait()

        with state.lock:
            state.process = None
        if process.returncode not in (0, None) and not state.stopped:
            had_failure = True
        if os.path.exists(group['output_path']):
            group_outputs.append(group['output_path'])

    _merge_xlsx_outputs(group_outputs, state.output_file)

    for group in state.pending_groups:
        for path in (group['input_path'], group['output_path']):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass

    with state.lock:
        if state.stopped:
            state.job_status = 'idle'
        elif had_failure:
            state.job_status = 'failed_unseen'
        else:
            state.job_status = 'completed_unseen'


@app.route('/api/scraper/analyze', methods=['POST'])
@login_required_api
def analyze_scraper_input():
    """Preview step: classifies the submitted URLs (file upload or pasted text)
    without starting anything, so the frontend can show a # | URL | Type |
    Scraper table and let the user confirm before any scraper runs.
    """
    try:
        raw_items = extract_input_source(request)
    except InputError as exc:
        return jsonify({'error': str(exc)}), 400

    entries, errors, unsupported = build_entries(raw_items)

    return jsonify({
        'entries': entries,
        'errors': errors,
        'unsupported': unsupported,
        'message': format_invalid_url_message(errors) or format_unsupported_message(unsupported) or None,
    })


@app.route('/StartScraper', methods=['POST'])
@login_required_api
def start_scraper():
    state = get_scraper_session()

    try:
        raw_items = extract_input_source(request)
    except InputError as exc:
        return jsonify({'error': str(exc)}), 400

    entries, errors, unsupported = build_entries(raw_items)

    if not entries:
        message = (
            format_invalid_url_message(errors)
            or format_unsupported_message(unsupported)
            or 'No valid URLs were provided.'
        )
        return jsonify({'error': message, 'errors': errors, 'unsupported': unsupported}), 400

    with state.lock:
        process_running = state.process is not None and state.process.poll() is None
        if process_running or state.job_status == 'running':
            return jsonify({'error': 'Scraper is already running.'}), 409

        state.url_statuses.clear()
        state.stopped = False
        state.job_id = uuid.uuid4().hex
        state.job_status = 'running'
        state.skipped = {'invalid': errors, 'unsupported': unsupported}

        job_id_short = state.job_id[:8]

        # Group URLs by detected type, in first-seen order, so each group can
        # be handed to exactly the one scraper script SCRIPT_MAP says handles
        # that type -- never a script name the request itself supplied.
        groups_by_type = OrderedDict()
        for entry in entries:
            groups_by_type.setdefault(entry['type'], []).append(entry['url'])

        os.makedirs(TMP_SCRAPERS_DIR, exist_ok=True)
        pending_groups = []
        for url_type, urls in groups_by_type.items():
            input_path = os.path.join(TMP_SCRAPERS_DIR, f'job_{job_id_short}_{url_type}.csv')
            with open(input_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for url in urls:
                    writer.writerow([url])

            pending_groups.append({
                'type': url_type,
                'input_path': input_path,
                'output_path': os.path.join(TMP_SCRAPERS_DIR, f'job_{job_id_short}_{url_type}_output.xlsx'),
                'script_path': os.path.join(BASE_DIR, 'scrapers', SCRIPT_MAP[url_type]),
            })
        state.pending_groups = pending_groups

        timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
        state.output_file = os.path.join(BASE_DIR, f'pitstoparabia_data_{job_id_short}_{timestamp}.xlsx')

        state.thread = threading.Thread(target=_run_job_groups, args=(state,), daemon=True)
        state.thread.start()

    return jsonify({
        'message': f'Scraper started. Job ID: {state.job_id}',
        'jobId': state.job_id,
        'groups': [g['type'] for g in pending_groups],
        'skipped': {'invalid': errors, 'unsupported': unsupported},
    })


@app.route('/stop-scraper', methods=['POST'])
@login_required_api
def stop_scraper():
    state = get_scraper_session()
    with state.lock:
        process_running = state.process is not None and state.process.poll() is None
        # job_status stays 'running' for the whole multi-group job, even in the
        # brief gap between one group's subprocess ending and the next one
        # starting (state.process is momentarily None then) -- so a stop
        # request landing in that gap must still take effect, not just be a
        # no-op that lets the next group start anyway.
        if not process_running and state.job_status != 'running':
            return jsonify({'stopped': False, 'message': 'No scraper process is running.'})

        state.stopped = True
        if process_running:
            try:
                state.process.terminate()
                state.process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                state.process.kill()
                state.process.wait()
            finally:
                state.process = None

    return jsonify({'stopped': True, 'message': 'Scraper has been stopped.'})

@app.route('/scraper-url-statuses')
@login_required_api
def scraper_url_statuses_endpoint():
    state = get_scraper_session()
    with state.lock:
        statuses = list(state.url_statuses)
        running = state.process is not None and state.process.poll() is None
        output_file = state.output_file
        # A completed/failed job's tree is served exactly once (the same request
        # cycle that /scraper-status reports it in), then cleared. Any later
        # fetch -- including a fresh page load after a refresh -- sees the
        # empty/idle tree instead of restoring a job that already finished.
        if not running and state.job_status == 'idle' and state.url_statuses:
            state.url_statuses.clear()

    xlsx_count, xlsx_urls = get_xlsx_info(output_file)
    if xlsx_urls:
        if statuses:
            for item in statuses:
                u = item.get('url', '').strip()
                if u in xlsx_urls:
                    item['status'] = 'done'
                    item['written_to_xlsx'] = True
        else:
            statuses = [
                {'url': u, 'status': 'done', 'parent': '', 'type': 'product', 'written_to_xlsx': True}
                for u in sorted(xlsx_urls)
            ]

    return jsonify({
        'statuses': statuses,
        'summary': build_status_summary(statuses),
        'xlsx_count': xlsx_count,
    })


@app.route('/scraper-status')
@login_required_api
def scraper_status():
    state = get_scraper_session()
    with state.lock:
        running = state.process is not None and state.process.poll() is None
        if running:
            state.job_status = 'running'

        current = state.job_status
        if current in ('completed_unseen', 'failed_unseen'):
            reported_status = 'completed' if current == 'completed_unseen' else 'failed'
            # One-shot: archive immediately so no later request (poll or page
            # refresh) ever sees this finished job as active again.
            state.job_status = 'idle'
        elif current == 'running':
            reported_status = 'running'
        else:
            reported_status = 'idle'

        has_active_job = reported_status == 'running'
        job_id = state.job_id if has_active_job else None
        output_file = state.output_file

    output_available = bool(output_file and os.path.exists(output_file))

    return jsonify({
        # Legacy fields kept for the existing frontend.
        'running': reported_status == 'running',
        'done': reported_status in ('completed', 'failed'),
        'outputAvailable': output_available,
        'outputFile': os.path.basename(output_file) if output_available else '',
        # New fields: explicit job status/identity.
        'status': reported_status,
        'hasActiveJob': has_active_job,
        'jobId': job_id,
    })


@app.route('/download-output')
@login_required_api
def download_output():
    state = get_scraper_session()
    with state.lock:
        output_file = state.output_file

    if not output_file or not os.path.exists(output_file):
        return Response('No output file found.', status=404, mimetype='text/plain')

    return send_from_directory(BASE_DIR, os.path.basename(output_file), as_attachment=True)


def _parse_urls_text(urls_text):
    """Returns (urls, error_message_or_None). Rejects the whole submission
    (rather than silently dropping bad rows) so the caller can point the
    error back at the URLs field for the user to fix.
    """
    raw_urls = parse_text_urls(urls_text)
    urls, errors = validate_url_list(raw_urls)
    if errors:
        lines = [f"Invalid URL on row {e['row']}: {e['value']}" for e in errors]
        return [], 'Invalid URL(s):\n' + '\n'.join(lines)
    if not urls:
        return [], 'At least one valid URL is required.'
    return urls, None


@app.route('/api/files/running')
@login_required_api
def api_list_running_files():
    """Every currently-running registered scraper (fileId + siteName only) --
    powers the Scraper page's switcher for hopping between several jobs
    started from /files without losing track of the others.
    """
    rows, _ = files_repo.list_files(per_page=200)
    running = [
        {'fileId': r['file_id'], 'siteName': r['site_name']}
        for r in rows
        if file_scraper_runner.is_running(r['file_id'])
    ]
    return jsonify({'files': running})


@app.route('/api/files')
@login_required_api
def api_list_files():
    search = request.args.get('search', '').strip() or None
    raw_trash = request.args.get('trash')
    is_deleted = None
    if raw_trash is not None:
        is_deleted = raw_trash.lower() in ('1', 'true', 'yes')

    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('perPage', 20))
    except ValueError:
        per_page = 20

    rows, total = files_repo.list_files(search=search, is_deleted=is_deleted, page=page, per_page=per_page)
    serialized = []
    user_id = session.get('user_id')
    active_map = job_manager.get_all_active_jobs_map()
    all_outputs = file_scraper_runner.get_all_output_paths()

    for r in rows:
        fid = r['file_id']
        item = files_repo.serialize_file(r)
        active = active_map.get(fid)
        if active:
            item['working'] = True
            item['is_owner'] = (active.get('user_id') == user_id)
        else:
            item['working'] = False
            item['is_owner'] = True
        item['outputAvailable'] = bool(all_outputs.get(fid))
        serialized.append(item)

    any_running = bool(active_map)
    has_any_output = bool(all_outputs)

    return jsonify({
        'files': serialized,
        'total': total,
        'page': page,
        'perPage': per_page,
        'anyRunning': any_running,
        'hasAnyOutput': has_any_output,
    })


@app.route('/api/files', methods=['POST'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
@require_csrf
def api_create_file():
    data = request.get_json(silent=True) or {}
    site_name = (data.get('siteName') or '').strip()
    python_file_path = (data.get('pythonFilePath') or '').strip()
    logo = (data.get('logo') or '').strip() or None
    urls_text = data.get('urlsText') or ''

    if not site_name:
        return jsonify({'error': 'Name is required.'}), 400
    if not python_file_path:
        return jsonify({'error': 'Python file is required.'}), 400

    urls, url_error = _parse_urls_text(urls_text)
    if url_error:
        return jsonify({'error': url_error}), 400

    try:
        created_by = session.get('user_id')
        file_id = files_repo.create_file(logo, site_name, python_file_path, created_by=created_by)
    except files_repo.FileValidationError as exc:
        return jsonify({'error': str(exc)}), 400

    files_repo.set_urls(file_id, urls)
    return jsonify({'file': files_repo.serialize_file(files_repo.get_file(file_id))}), 201


@app.route('/api/files/<int:file_id>', methods=['PUT'])
@login_required_api
@require_csrf
def api_update_file(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404
    if file_scraper_runner.is_running(file_id):
        return jsonify({'error': 'Stop this scraper before editing it.'}), 409

    data = request.get_json(silent=True) or {}
    site_name = (data.get('siteName') or '').strip()
    python_file_path = (data.get('pythonFilePath') or '').strip()
    logo = (data.get('logo') or '').strip() or None
    urls_text = data.get('urlsText') or ''

    if not site_name:
        return jsonify({'error': 'Name is required.'}), 400
    if not python_file_path:
        return jsonify({'error': 'Python file is required.'}), 400

    urls, url_error = _parse_urls_text(urls_text)
    if url_error:
        return jsonify({'error': url_error}), 400

    try:
        files_repo.update_file(file_id, logo, site_name, python_file_path)
    except files_repo.FileValidationError as exc:
        return jsonify({'error': str(exc)}), 400

    files_repo.set_urls(file_id, urls)
    return jsonify({'file': files_repo.serialize_file(files_repo.get_file(file_id))})


@app.route('/api/files/<int:file_id>', methods=['DELETE'])
@login_required_api
@require_csrf
def api_delete_file(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404
    if file_scraper_runner.is_running(file_id):
        return jsonify({'error': 'Stop this scraper before deleting it.'}), 409

    files_repo.delete_file(file_id)
    return jsonify({'message': 'Scraper permanently deleted.'})


@app.route('/api/files/<int:file_id>/toggle-status', methods=['POST'])
@login_required_api
@require_csrf
def api_toggle_file_status(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404

    if file_scraper_runner.is_running(file_id):
        return jsonify({'error': 'Stop this scraper before disabling it.'}), 409

    data = request.get_json(silent=True) or {}
    currently_deleted = files_repo.bit_to_bool(record.get('is_deleted'))

    if 'enabled' in data:
        new_enabled = bool(data['enabled'])
    else:
        new_enabled = currently_deleted  # toggle

    files_repo.set_file_enabled(file_id, new_enabled)
    updated = files_repo.get_file(file_id)
    return jsonify({
        'success': True,
        'isEnabled': new_enabled,
        'isDeleted': not new_enabled,
        'message': 'Scraper enabled.' if new_enabled else 'Scraper disabled.',
        'file': files_repo.serialize_file(updated),
    })


@app.route('/api/files/<int:file_id>/restore', methods=['POST'])
@login_required_api
@require_csrf
def api_restore_file(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404

    files_repo.restore_file(file_id)
    updated = files_repo.get_file(file_id)
    return jsonify({
        'success': True,
        'message': 'Scraper restored to Active list.',
        'file': files_repo.serialize_file(updated),
    })


# ==============================================================================
# Centralized Job-Based Scraper APIs
# ==============================================================================

# ==============================================================================
# Centralized Job-Based Scraper APIs (with User Ownership Protection)
# ==============================================================================

@app.route('/api/scraper/start', methods=['POST'])
@login_required_api
@require_csrf
def api_scraper_job_start():
    data = request.get_json(silent=True) or {}
    file_id = data.get('file_id') or request.form.get('file_id') or request.args.get('file_id')
    if not file_id:
        return jsonify({'success': False, 'error': 'Missing file_id parameter.'}), 400
    try:
        file_id = int(file_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid file_id.'}), 400

    user_id = session.get('user_id')
    result = job_manager.start_job(file_id, user_id=user_id)
    if not result.get('success'):
        return jsonify(result), 409
    return jsonify(result)


@app.route('/api/scraper/file/<int:file_id>/active-job')
@login_required_api
def api_scraper_file_active_job(file_id):
    user_id = session.get('user_id')
    active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)
    return jsonify(active_info)


@app.route('/api/scraper/job/<string:job_id>/status')
@login_required_api
def api_scraper_job_status(job_id):
    user_id = session.get('user_id')
    data, code = job_manager.get_job_status(job_id, current_user_id=user_id)
    return jsonify(data), code


@app.route('/api/scraper/job/<string:job_id>/urls')
@login_required_api
def api_scraper_job_urls(job_id):
    user_id = session.get('user_id')
    urls, code = job_manager.get_job_urls(job_id, current_user_id=user_id)
    if code != 200:
        return jsonify(urls), code
    summary = build_status_summary(urls)
    return jsonify({
        'job_id': job_id,
        'statuses': urls,
        'summary': summary,
        'count': len(urls),
    })


@app.route('/api/scraper/job/<string:job_id>/events')
@login_required_api
def api_scraper_job_events(job_id):
    """Server-Sent Events (SSE) Webhook endpoint for live scraper progress and URL updates."""
    user_id = session.get('user_id')
    job = job_manager.get_log_by_job_id(job_id)
    if not job:
        with job_manager._lock:
            state = job_manager._active_jobs.get(job_id)
        if not state:
            return jsonify({'error': 'Job not found.'}), 404
        if state['started_by_user_id'] != user_id and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden'}), 403
    elif job['user_id'] != user_id and session.get('role') != 'SuperAdmin':
        return jsonify({'error': 'Forbidden'}), 403

    import queue

    def event_stream():
        q = job_manager.subscribe_sse(job_id)
        try:
            # 1. Send initial snapshot immediately
            status_data, _ = job_manager.get_job_status(job_id, current_user_id=user_id)
            urls_data, _ = job_manager.get_job_urls(job_id, current_user_id=user_id)
            initial_payload = {
                'type': 'snapshot',
                'summary': status_data,
                'statuses': urls_data
            }
            yield f"data: {json.dumps(initial_payload)}\n\n"

            # 2. Listen to queue and stream events
            while True:
                try:
                    event = q.get(timeout=15.0)
                    yield f"data: {json.dumps(event)}\n\n"
                    if event.get('done') or (event.get('type') == 'status' and event.get('status') in ('SUCCESS', 'STOPPED', 'FAILED', 'FAIL')):
                        break
                except queue.Empty:
                    # Heartbeat ping
                    yield ": ping\n\n"
        except GeneratorExit:
            pass
        finally:
            job_manager.unsubscribe_sse(job_id, q)

    return Response(
        stream_with_context(event_stream()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        }
    )


@app.route('/api/scraper/job/<string:job_id>/stop', methods=['POST'])
@login_required_api
@require_csrf
def api_scraper_job_stop(job_id):
    user_id = session.get('user_id')
    is_superadmin = (session.get('role') == 'SuperAdmin')
    result, code = job_manager.stop_job(job_id, current_user_id=user_id, is_superadmin=is_superadmin)
    return jsonify(result), code


# --- Backward Compatible Routes for /api/files/ endpoints ---

@app.route('/api/files/<int:file_id>/start', methods=['POST'])
@login_required_api
@require_csrf
def api_start_file(file_id):
    user_id = session.get('user_id')
    result = job_manager.start_job(file_id, user_id=user_id)
    if not result.get('success'):
        return jsonify(result), 409
    return jsonify(result)


@app.route('/api/files/<int:file_id>/stop', methods=['POST'])
@login_required_api
@require_csrf
def api_stop_file(file_id):
    user_id = session.get('user_id')
    is_superadmin = (session.get('role') == 'SuperAdmin')
    result, code = job_manager.stop_file(file_id, current_user_id=user_id, is_superadmin=is_superadmin)
    return jsonify(result), code


@app.route('/api/files/<int:file_id>/status')
@login_required_api
def api_file_status(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404

    user_id = session.get('user_id')
    active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)

    if active_info.get('job_id'):
        job_status, code = job_manager.get_job_status(active_info['job_id'], current_user_id=user_id)
        if code == 200:
            job_status['is_owner'] = True
            job_status['working'] = bool(active_info.get('has_active_job'))
            job_status['siteName'] = record['site_name']
            job_status['fileId'] = file_id
            return jsonify(job_status)

    output_path = file_scraper_runner.get_output_path(file_id)
    return jsonify({
        'job_id': None,
        'running': False,
        'working': False,
        'done': True,
        'is_owner': True,
        'siteName': record['site_name'],
        'fileId': file_id,
        'outputAvailable': bool(output_path and os.path.exists(output_path)),
        'total_product_urls': 0,
        'written_to_xlsx': 0,
        'pending': 0,
        'running_count': 0,
        'blocked': 0,
        'main_url_done': 0,
        'product_url_done': 0,
        'progress_percent': 0.0,
    })


@app.route('/api/files/<int:file_id>/url-statuses')
@login_required_api
def api_file_url_statuses(file_id):
    user_id = session.get('user_id')
    active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)

    if active_info['has_active_job'] and not active_info['is_owner']:
        return jsonify({'statuses': [], 'summary': {}, 'xlsx_count': 0, 'error': 'Forbidden'}), 403

    if active_info.get('job_id') and active_info.get('is_owner', True):
        urls, code = job_manager.get_job_urls(active_info['job_id'], current_user_id=user_id)
        if code == 200 and urls:
            summary = build_status_summary(urls)
            return jsonify({
                'statuses': urls,
                'summary': summary,
                'xlsx_count': summary.get('written_to_xlsx', 0),
            })

    urls = file_scraper_runner.get_statuses(file_id)
    summary = build_status_summary(urls)
    return jsonify({
        'statuses': urls,
        'summary': summary,
        'xlsx_count': summary.get('written_to_xlsx', 0),
    })


def _format_scraper_output_filename(site_name, extension='xlsx'):
    clean_site = re.sub(r'[^A-Za-z0-9]+', '_', site_name or '').strip('_').upper() or 'SCRAPER'
    today = datetime.now().strftime('%d-%m-%Y')
    return f"{clean_site}_{today}.{extension}"


@app.route('/api/files/<int:file_id>/download')
@login_required_api
def api_file_download(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404

    output_path = file_scraper_runner.get_output_path(file_id)
    if not output_path or not os.path.exists(output_path):
        return jsonify({'error': 'No output available for this scraper yet. Run it first.'}), 404

    filename = _format_scraper_output_filename(record['site_name'])
    return send_file(
        output_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/api/files/download-zip')
@login_required_api
def api_files_download_zip():
    # 1. Reject download if any crawler is still actively running
    if file_scraper_runner.running_count() > 0:
        return jsonify({
            'error': 'Scraping is in progress. Please wait until all scrapers finish before downloading ZIP.',
            'anyRunning': True,
        }), 409

    all_outputs = file_scraper_runner.get_all_output_paths()
    if not all_outputs:
        return jsonify({'error': 'No completed scraper reports available to download.'}), 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        used_names = set()
        for file_id, file_path in all_outputs.items():
            if not os.path.exists(file_path):
                continue
            record = files_repo.get_file(file_id)
            site_name = record['site_name'] if record else f"scraper_{file_id}"
            base_filename = _format_scraper_output_filename(site_name)
            filename = base_filename
            counter = 1
            while filename in used_names:
                root, ext = os.path.splitext(base_filename)
                filename = f"{root}_{counter}{ext}"
                counter += 1
            used_names.add(filename)
            zf.write(file_path, arcname=filename)

    if not used_names:
        return jsonify({'error': 'No valid output files found to package.'}), 404

    zip_buffer.seek(0)
    today = datetime.now().strftime('%d-%m-%Y')
    zip_filename = f"scrapers_output_{today}.zip"
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename,
    )


@app.route('/api/files/upload-script', methods=['POST'])
@login_required_api
@role_required_api('SuperAdmin', 'Admin')
@require_csrf
def api_upload_file_script():
    if not request.files or 'file' not in request.files:
        return jsonify({'error': 'No file uploaded.'}), 400

    upload = request.files['file']
    candidate_name = (upload.filename or '').strip()
    if candidate_name:
        existing = files_repo.get_file_by_path(candidate_name)
        if existing and file_scraper_runner.is_running(existing['file_id']):
            return jsonify({'error': 'Stop this scraper before replacing its Python file.'}), 409

    try:
        filename = files_repo.save_uploaded_script(upload)
    except files_repo.FileValidationError as exc:
        return jsonify({'error': str(exc)}), 400

    return jsonify({'fileName': filename})


@app.route('/api/files/parse-urls', methods=['POST'])
@login_required_api
def api_parse_urls():
    if request.files and 'file' in request.files:
        upload = request.files['file']
        filename = (upload.filename or '').lower()
        if not filename.endswith('.csv'):
            return jsonify({'error': 'Please upload a .csv file.'}), 400
        raw_bytes = upload.read()
        if not raw_bytes:
            return jsonify({'error': 'The uploaded file is empty.'}), 400
        raw_urls = [url for url, _declared_type in parse_csv_urls(raw_bytes)]
    else:
        data = request.get_json(silent=True) or {}
        raw_urls = parse_text_urls(data.get('text') or '')

    urls, errors = validate_url_list(raw_urls)
    if not urls:
        return jsonify({'error': 'No valid URLs were found.', 'errors': errors}), 400

    return jsonify({'urls': urls, 'errors': errors})


@app.route('/api/files/<int:file_id>/logs')
@login_required_api
def api_file_logs(file_id):
    record = files_repo.get_file(file_id)
    if not record:
        return jsonify({'error': 'Scraper not found.'}), 404
    rows, total = reports_repo.list_logs(file_id=file_id, per_page=100)
    return jsonify({
        'fileId': file_id,
        'siteName': record.get('site_name') or 'Scraper',
        'logs': [reports_repo.serialize_log(r) for r in rows],
        'total': total,
    })


@app.route('/api/reports')
@login_required_api
@role_required_api('SuperAdmin')
def api_list_reports():
    search = request.args.get('search', '').strip() or None
    status = request.args.get('status', '').strip() or None
    user_id_raw = request.args.get('userId', '').strip()
    user_id = int(user_id_raw) if user_id_raw.isdigit() else None
    file_id_raw = request.args.get('fileId', '').strip()
    file_id = int(file_id_raw) if file_id_raw.isdigit() else None
    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('perPage', 20))
    except ValueError:
        per_page = 20

    rows, total = reports_repo.list_logs(
        search=search,
        status=status,
        user_id=user_id,
        file_id=file_id,
        page=page,
        per_page=per_page,
    )
    stats = reports_repo.get_logs_summary_stats()
    return jsonify({
        'reports': [reports_repo.serialize_log(r) for r in rows],
        'total': total,
        'page': page,
        'perPage': per_page,
        'stats': stats,
    })


@app.route('/api/reports/<int:report_id>/download')
@login_required_api
@role_required_api('SuperAdmin')
def api_download_report_output(report_id):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM logTbl WHERE id = %s", (report_id,))
            record = cursor.fetchone()
    finally:
        conn.close()

    if not record or not record.get('output_file_path') or not os.path.exists(record['output_file_path']):
        return jsonify({'error': 'Output file not available for this run.'}), 404

    filename = _format_scraper_output_filename(record['scraper'])
    return send_file(
        record['output_file_path'],
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.errorhandler(404)
def handle_404_error(e):
    """Gracefully handles unwanted page or API requests by serving custom 404."""
    if request.path.startswith('/api/') or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'error': 'The requested API resource was not found.',
            'status': 404,
            'path': request.path
        }), 404
    return render_template(
        '404.html',
        page='404',
        requested_path=request.path,
        user_name=session.get('name'),
        user_email=session.get('email'),
        user_role=session.get('role'),
        user_avatar=session.get('avatar'),
        unread_notifications=0,
        notifications=[]
    ), 404


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
