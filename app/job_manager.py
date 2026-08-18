"""Centralized Scraper Execution Engine using logTbl as single source of truth.

Architectural Guarantees:
1. GLOBAL SCRAPER LOCK (in-memory + logTbl status check):
   - _start_lock threading.Lock prevents concurrent starts within this Flask process.
   - logTbl rows with status='RUNNING' and end_time IS NULL are the persisted lock.
2. USER-OWNED EXECUTION & PRIVACY:
   - Only the owner (user_id in logTbl) can view live counters, progress, URLs, and Stop button.
   - Non-owners receive ONLY:
     {"already_running": true, "is_owner": false, "message": "This scraper is currently being used by another user."}
   - No job_id, process_id, logs, timestamps, or usernames are ever leaked to non-owners.
3. SSE (Server-Sent Events / Webhook):
   - Each active job has a thread-safe queue in _sse_queues[job_id].
   - subscribe_sse(job_id) / unsubscribe_sse(job_id, q) manage client connections.
4. LOCK RELEASE & LIFECYCLE:
   - finalize_job() marks end_time + terminal status in logTbl.
   - 6-Hour Timeout Watchdog ensures dead jobs never hold the lock.
"""

import csv
import json
import logging
import os
import queue
import subprocess
import sys
import threading
import time
import uuid
from collections import OrderedDict
from datetime import datetime, timezone

import psutil

from db import get_connection
import files_repo
import reports_repo
from scraper_status_utils import parse_status_line

logger = logging.getLogger(__name__)

BASE_DIR = files_repo.BASE_DIR
TMP_DIR = os.path.join(BASE_DIR, 'tmp', 'file_scrapers')
os.makedirs(TMP_DIR, exist_ok=True)

MAX_RUNTIME_SECONDS = 6 * 3600

_lock = threading.RLock()
_start_lock = threading.Lock()
_active_jobs = {}
_sse_queues = {}
_sse_lock = threading.Lock()


def _push_sse_event(job_id, event_data):
    with _sse_lock:
        for q in _sse_queues.get(job_id, []):
            try:
                q.put_nowait(event_data)
            except Exception:
                pass


def subscribe_sse(job_id):
    q = queue.Queue(maxsize=500)
    with _sse_lock:
        _sse_queues.setdefault(job_id, []).append(q)
    return q


def unsubscribe_sse(job_id, q):
    with _sse_lock:
        lst = _sse_queues.get(job_id, [])
        if q in lst:
            lst.remove(q)
        if not lst and job_id in _sse_queues:
            del _sse_queues[job_id]


def _popen_kwargs():
    if os.name == 'nt':
        return {'creationflags': subprocess.CREATE_NEW_PROCESS_GROUP}
    return {'start_new_session': True}


def _kill_process_tree(process, force=False):
    if not process:
        return
    pid = process.pid if hasattr(process, 'pid') else process
    if not pid:
        return
    try:
        if psutil.pid_exists(pid):
            parent = psutil.Process(pid)
            for child in parent.children(recursive=True):
                try:
                    child.kill() if force else child.terminate()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            try:
                parent.kill() if force else parent.terminate()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
            return
    except Exception:
        pass
    if hasattr(process, 'terminate'):
        try:
            process.kill() if force else process.terminate()
        except (ProcessLookupError, OSError):
            pass


def _cleanup_temp(*paths):
    for path in paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


def get_active_log(file_id):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, job_id, scraper, file_id, user_id, status,
                       no_of_url_found, total_success_url, total_block_url, data_scraped,
                       total_products, pending_urls, running_urls, completed_urls, blocked_urls,
                       main_url_done, product_url_done, progress_percent,
                       output_file_path, error_message, process_id,
                       start_time, end_time
                FROM logTbl
                WHERE file_id = %s AND status = 'RUNNING' AND end_time IS NULL
                ORDER BY id DESC LIMIT 1
            """, (file_id,))
            return cursor.fetchone()
    finally:
        conn.close()


def get_active_job(file_id):
    """Alias for get_active_log to maintain backwards compatibility."""
    return get_active_log(file_id)


def get_log_by_job_id(job_id):
    if not job_id:
        return None
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, job_id, scraper, file_id, user_id, status,
                       no_of_url_found, total_success_url, total_block_url, data_scraped,
                       total_products, pending_urls, running_urls, completed_urls, blocked_urls,
                       main_url_done, product_url_done, progress_percent,
                       output_file_path, error_message, process_id,
                       start_time, end_time
                FROM logTbl WHERE job_id = %s LIMIT 1
            """, (job_id,))
            return cursor.fetchone()
    finally:
        conn.close()


def get_job_by_id(job_id):
    """Alias for get_log_by_job_id for backwards compatibility."""
    return get_log_by_job_id(job_id)


def _close_stale_log_row(log_id):
    try:
        conn = get_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "UPDATE logTbl SET status='FAIL', end_time=NOW(), "
                "error_message='Process terminated unexpectedly.' WHERE id=%s",
                (log_id,)
            )
        conn.close()
    except Exception:
        pass


def finalize_job(job_id, status, error_message=None, output_file_path=None, final_counters=None):
    log_status = 'SUCCESS' if status == 'SUCCESS' else ('STOPPED' if status == 'STOPPED' else 'FAIL')
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            if final_counters:
                cursor.execute("""
                    UPDATE logTbl SET
                        status=%s, end_time=NOW(),
                        error_message=COALESCE(%s, error_message),
                        output_file_path=COALESCE(%s, output_file_path),
                        no_of_url_found=%s, total_success_url=%s, total_block_url=%s,
                        data_scraped=%s, total_products=%s, pending_urls=0, running_urls=0,
                        completed_urls=%s, blocked_urls=%s, main_url_done=%s,
                        product_url_done=%s, progress_percent=%s
                    WHERE job_id=%s
                """, (
                    log_status, error_message, output_file_path,
                    final_counters.get('total_urls', 0),
                    final_counters.get('completed', 0),
                    final_counters.get('blocked', 0),
                    final_counters.get('written_to_xlsx', 0),
                    final_counters.get('total_products', 0),
                    final_counters.get('completed', 0),
                    final_counters.get('blocked', 0),
                    final_counters.get('main_url_done', 0),
                    final_counters.get('product_url_done', 0),
                    final_counters.get('progress_percent', 0.0),
                    job_id
                ))
            else:
                cursor.execute("""
                    UPDATE logTbl SET status=%s, end_time=NOW(),
                    error_message=COALESCE(%s, error_message),
                    output_file_path=COALESCE(%s, output_file_path)
                    WHERE job_id=%s
                """, (log_status, error_message, output_file_path, job_id))
    except Exception:
        logger.exception('Error finalizing job_id=%s', job_id)
    finally:
        conn.close()
    _push_sse_event(job_id, {
        'type': 'status', 'status': log_status, 'done': True, 'error_message': error_message
    })


def _recalc_counters(job_state):
    urls_dict = job_state.get('urls', {})
    items = list(urls_dict.values())
    main_urls = [u for u in items if u.get('type') == 'sitemap' or 'sitemap' in u.get('url', '').lower()]
    product_urls = [u for u in items if u not in main_urls]
    main_url_done = sum(1 for u in main_urls if (u.get('status') or '').lower() in ('done', 'success'))
    product_url_done = sum(1 for u in product_urls if (u.get('status') or '').lower() in ('done', 'success'))
    total_products = len(product_urls)
    total_urls = len(items)
    pending = sum(1 for u in items if (u.get('status') or '').lower() == 'pending')
    running = sum(1 for u in items if (u.get('status') or '').lower() == 'running')
    blocked = sum(1 for u in items if (u.get('status') or '').lower() in ('blocked', 'failed'))
    completed = main_url_done + product_url_done
    written_to_xlsx = product_url_done
    output_path = job_state.get('output_file_path')
    if output_path and os.path.exists(output_path):
        excel_rows = reports_repo.count_excel_data_rows(output_path)
        if excel_rows > 0:
            written_to_xlsx = excel_rows
    if total_products > 0:
        progress_pct = round(min(100.0, (product_url_done / float(total_products)) * 100.0), 1)
    elif total_urls > 0:
        progress_pct = round(min(100.0, (completed / float(total_urls)) * 100.0), 1)
    else:
        progress_pct = 0.0
    job_state.update({
        'main_url_done': main_url_done, 'product_url_done': product_url_done,
        'total_products': total_products, 'total_urls': total_urls,
        'pending': pending, 'running': running, 'blocked': blocked,
        'completed': completed, 'written_to_xlsx': written_to_xlsx,
        'progress_percent': progress_pct
    })


def _build_sse_status(state):
    return {
        'type': 'status',
        'job_id': state.get('job_id'),
        'file_id': state.get('file_id'),
        'site_name': state.get('site_name', 'Scraper'),
        'status': state.get('status', 'RUNNING'),
        'running': state.get('status') == 'RUNNING',
        'done': state.get('status') in ('SUCCESS', 'FAILED', 'STOPPED', 'FAIL'),
        'total_product_urls': state.get('total_products', 0),
        'total_urls': state.get('total_urls', 0),
        'written_to_xlsx': state.get('written_to_xlsx', 0),
        'pending': state.get('pending', 0),
        'running_count': state.get('running', 0),
        'blocked': state.get('blocked', 0),
        'completed': state.get('completed', 0),
        'main_url_done': state.get('main_url_done', 0),
        'product_url_done': state.get('product_url_done', 0),
        'progress_percent': state.get('progress_percent', 0.0),
        'output_available': bool(
            state.get('output_file_path') and os.path.exists(state['output_file_path'])
            and state.get('status') != 'RUNNING'
        ),
        'error_message': state.get('error_message'),
    }


def _sync_counters_to_db(job_id, state):
    try:
        conn = get_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE logTbl SET
                    no_of_url_found=%s, total_success_url=%s, total_block_url=%s, data_scraped=%s,
                    total_products=%s, pending_urls=%s, running_urls=%s, completed_urls=%s,
                    blocked_urls=%s, main_url_done=%s, product_url_done=%s, progress_percent=%s
                WHERE job_id=%s
            """, (
                state['total_urls'], state['completed'], state['blocked'], state['written_to_xlsx'],
                state['total_products'], state['pending'], state['running'], state['completed'],
                state['blocked'], state['main_url_done'], state['product_url_done'],
                state['progress_percent'], job_id
            ))
        conn.close()
    except Exception:
        logger.exception('Counter sync failed for job_id=%s', job_id)


def _run_worker(job_id, file_id, script_path, input_path, output_placeholder):
    with _lock:
        job_state = _active_jobs.get(job_id)
        if not job_state:
            return
        process = job_state['process']

    returncode = -1
    try:
        for line in iter(process.stdout.readline, ''):
            if not line:
                break
            parsed = parse_status_line(line.rstrip('\n'))
            if parsed:
                with _lock:
                    if job_id in _active_jobs:
                        state = _active_jobs[job_id]
                        url_key = parsed['url']
                        existing = state['urls'].get(url_key)
                        if existing:
                            existing['status'] = parsed['status']
                            if parsed.get('parent'):
                                existing['parent'] = parsed['parent']
                            if parsed.get('type'):
                                existing['type'] = parsed['type']
                        else:
                            state['urls'][url_key] = {
                                'url': parsed['url'],
                                'status': parsed['status'],
                                'parent': parsed.get('parent') or '',
                                'type': parsed.get('type') or 'root',
                            }
                        _recalc_counters(state)
                        now = time.time()
                        if now - state.get('last_db_sync', 0) >= 3.0:
                            state['last_db_sync'] = now
                            _sync_counters_to_db(job_id, state)
                        # Push live SSE event
                        _push_sse_event(job_id, {
                            'type': 'url_update',
                            'url': state['urls'][url_key],
                            'summary': _build_sse_status(state),
                        })
        process.stdout.close()
        process.wait()
        returncode = process.returncode
    except Exception:
        logger.exception('Scraper job_id=%s crashed during monitoring.', job_id)
    finally:
        with _lock:
            state = _active_jobs.get(job_id, {})
            was_stopped = state.get('stopped', False)
            was_timeout = state.get('timeout_stopped', False)
            _recalc_counters(state)

        final_output_path = output_placeholder if (output_placeholder and os.path.exists(output_placeholder)) else None
        total_blocked = state.get('blocked', 0)
        total_completed = state.get('completed', 0)
        written_count = state.get('written_to_xlsx', 0)

        if was_stopped:
            final_status = 'STOPPED'
            error_message = 'Scraper stopped by user.'
        elif was_timeout:
            final_status = 'STOPPED'
            error_message = 'Automatically stopped after maximum runtime of 6 hours.'
        elif total_blocked > 0 and total_completed == 0:
            final_status = 'FAILED'
            error_message = f'Scraping failed: {total_blocked} URLs blocked by target website.'
        elif returncode == 0:
            final_status = 'SUCCESS'
            error_message = None
        else:
            final_status = 'FAILED'
            error_message = f'Scraper process exited with return code {returncode}.'

        finalize_job(
            job_id=job_id, status=final_status,
            error_message=error_message, output_file_path=final_output_path,
            final_counters={
                'total_urls': state.get('total_urls', 0),
                'pending': 0, 'completed': total_completed, 'blocked': total_blocked,
                'total_products': state.get('total_products', 0),
                'written_to_xlsx': written_count,
                'main_url_done': state.get('main_url_done', 0),
                'product_url_done': state.get('product_url_done', 0),
                'progress_percent': state.get('progress_percent', 0.0),
            }
        )
        files_repo.set_working(file_id, 0)
        _cleanup_temp(input_path)
        with _lock:
            if job_id in _active_jobs:
                _active_jobs[job_id]['status'] = final_status
                _active_jobs[job_id]['running'] = 0


def start_job(file_id, user_id):
    record = files_repo.get_file(file_id)
    if not record:
        return {'success': False, 'error': 'Scraper not found.'}
    if files_repo.bit_to_bool(record.get('is_deleted')):
        return {'success': False, 'error': 'This scraper is disabled. Please enable it first.'}

    with _start_lock:
        active = get_active_log(file_id)
        if active:
            job_id = active.get('job_id')
            with _lock:
                in_memory = (job_id in _active_jobs) if job_id else False
            pid = active.get('process_id')
            is_alive = in_memory or (pid and psutil.pid_exists(pid))
            if is_alive:
                if active['user_id'] == user_id:
                    return {'success': True, 'job_id': job_id, 'file_id': file_id,
                            'is_owner': True, 'is_new_job': False,
                            'message': 'Resuming your active scraper execution.'}
                else:
                    return {'success': False, 'already_running': True, 'is_owner': False,
                            'message': 'This scraper is currently being used by another user.'}
            else:
                if job_id:
                    finalize_job(job_id, status='FAILED', error_message='Process terminated unexpectedly.')
                else:
                    _close_stale_log_row(active['id'])
                files_repo.set_working(file_id, 0)

        new_job_id = uuid.uuid4().hex[:16]
        try:
            script_path = files_repo.resolve_script_path(record['python_file_path'])
        except Exception as exc:
            return {'success': False, 'error': str(exc)}

        urls = []
        if record.get('urls_json'):
            try:
                urls = json.loads(record['urls_json'])
            except Exception:
                urls = []

        output_placeholder = os.path.join(TMP_DIR, f'file_{file_id}_output.xlsx')
        input_path = None
        if urls:
            input_path = os.path.join(TMP_DIR, f'file_{file_id}_urls.csv')
            with open(input_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for u in urls:
                    writer.writerow([u])

        args = [sys.executable, '-u', script_path, output_placeholder]
        if input_path:
            args.append(input_path)

        try:
            process = subprocess.Popen(
                args, cwd=BASE_DIR, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, **_popen_kwargs(),
            )
        except OSError as exc:
            _cleanup_temp(input_path, output_placeholder)
            return {'success': False, 'error': f'Failed to launch process: {exc}'}

        scraper_name = record.get('site_name') or 'Scraper'
        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO logTbl
                        (job_id, scraper, file_id, user_id, status, start_time,
                         process_id, output_file_path, total_products, pending_urls)
                    VALUES (%s, %s, %s, %s, 'RUNNING', NOW(), %s, %s, %s, %s)
                """, (new_job_id, scraper_name, file_id, user_id,
                      process.pid, output_placeholder, len(urls), len(urls)))
        except Exception:
            _cleanup_temp(input_path, output_placeholder)
            logger.exception('Failed to insert logTbl row for file_id=%s', file_id)
            return {'success': False, 'error': 'Failed to record job start.'}
        finally:
            conn.close()

        files_repo.set_working(file_id, 1)
        initial_urls = OrderedDict()
        for u in urls:
            initial_urls[u] = {'url': u, 'status': 'pending', 'parent': '', 'type': 'root'}

        with _lock:
            _active_jobs[new_job_id] = {
                'job_id': new_job_id, 'file_id': file_id,
                'started_by_user_id': user_id, 'site_name': scraper_name,
                'process': process, 'pid': process.pid,
                'started_at': datetime.now(timezone.utc), 'status': 'RUNNING',
                'urls': initial_urls, 'total_products': len(urls), 'total_urls': len(urls),
                'written_to_xlsx': 0, 'pending': len(urls), 'running': 0, 'blocked': 0,
                'completed': 0, 'main_url_done': 0, 'product_url_done': 0,
                'progress_percent': 0.0, 'output_file_path': output_placeholder,
                'error_message': None, 'stopped': False, 'timeout_stopped': False,
                'last_db_sync': time.time(),
            }

        threading.Thread(
            target=_run_worker,
            args=(new_job_id, file_id, script_path, input_path, output_placeholder),
            daemon=True,
        ).start()

        return {'success': True, 'job_id': new_job_id, 'file_id': file_id,
                'is_owner': True, 'is_new_job': True,
                'message': 'Scraper job started successfully.'}


def get_active_job_for_file(file_id, current_user_id):
    active = get_active_log(file_id)
    record = files_repo.get_file(file_id)
    site_name = record['site_name'] if record else 'Scraper'
    if not active:
        return {'has_active_job': False, 'already_running': False, 'is_owner': True,
                'file_id': file_id, 'site_name': site_name, 'status': 'IDLE'}
    job_id = active.get('job_id')
    with _lock:
        in_memory = (job_id in _active_jobs) if job_id else False
    pid = active.get('process_id')
    is_alive = in_memory or (pid and psutil.pid_exists(pid))
    if not is_alive:
        if job_id:
            finalize_job(job_id, status='FAILED', error_message='Process terminated unexpectedly.')
        else:
            _close_stale_log_row(active['id'])
        files_repo.set_working(file_id, 0)
        return {'has_active_job': False, 'already_running': False, 'is_owner': True,
                'file_id': file_id, 'site_name': site_name, 'status': 'IDLE'}
    is_owner = (active['user_id'] == current_user_id)
    if not is_owner:
        return {'has_active_job': True, 'already_running': True, 'is_owner': False,
                'file_id': file_id, 'site_name': site_name, 'status': 'RUNNING',
                'message': 'This scraper is currently being used by another user.'}
    return {'has_active_job': True, 'already_running': True, 'is_owner': True,
            'job_id': job_id, 'file_id': file_id, 'site_name': site_name, 'status': active['status']}


def get_job_status(job_id, current_user_id):
    with _lock:
        state = _active_jobs.get(job_id)
        if state:
            if state['started_by_user_id'] != current_user_id:
                return {'success': False, 'error': 'Forbidden'}, 403
            _recalc_counters(state)
            return _build_sse_status(state), 200
    job = get_log_by_job_id(job_id)
    if not job:
        return {'success': False, 'error': 'Job not found.'}, 404
    if job['user_id'] != current_user_id:
        return {'success': False, 'error': 'Forbidden'}, 403
    is_running = (job['status'] == 'RUNNING' and job['end_time'] is None)
    output_path = job.get('output_file_path')
    record = files_repo.get_file(job['file_id'])
    return {
        'job_id': job['job_id'], 'file_id': job['file_id'],
        'site_name': record['site_name'] if record else 'Scraper',
        'status': job['status'], 'running': is_running, 'done': not is_running,
        'total_product_urls': job.get('total_products') or 0,
        'total_urls': job.get('no_of_url_found') or 0,
        'written_to_xlsx': job.get('data_scraped') or 0,
        'pending': job.get('pending_urls') or 0,
        'running_count': job.get('running_urls') or 0,
        'blocked': job.get('blocked_urls') or 0,
        'completed': job.get('completed_urls') or 0,
        'main_url_done': job.get('main_url_done') or 0,
        'product_url_done': job.get('product_url_done') or 0,
        'progress_percent': job.get('progress_percent') or 0.0,
        'started_at': job['start_time'].isoformat() if job.get('start_time') else None,
        'output_available': bool(output_path and os.path.exists(output_path) and not is_running),
        'output_file_path': output_path, 'error_message': job.get('error_message'),
    }, 200


def get_job_urls(job_id, current_user_id):
    with _lock:
        state = _active_jobs.get(job_id)
        if state:
            if state['started_by_user_id'] != current_user_id:
                return {'success': False, 'error': 'Forbidden'}, 403
            return list(state['urls'].values()), 200
    job = get_log_by_job_id(job_id)
    if not job:
        return {'success': False, 'error': 'Job not found.'}, 404
    if job['user_id'] != current_user_id:
        return {'success': False, 'error': 'Forbidden'}, 403
    if job.get('output_file_path') and os.path.exists(job['output_file_path']):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(job['output_file_path'], read_only=True, data_only=True)
            sheet = wb.active
            headers = [str(c).strip().lower() for c in next(sheet.iter_rows(values_only=True), [])]
            url_col_idx = next((i for i, h in enumerate(headers) if h in ('url', 'source', 'product url', 'link')), None)
            urls = []
            if url_col_idx is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > url_col_idx and row[url_col_idx]:
                        urls.append({'url': str(row[url_col_idx]).strip(), 'status': 'done', 'parent': '', 'type': 'product'})
            wb.close()
            return urls, 200
        except Exception:
            pass
    return [], 200


def stop_job(job_id, current_user_id, is_superadmin=False):
    job = get_log_by_job_id(job_id)
    if not job:
        return {'success': False, 'error': 'Job not found.'}, 404
    if job['user_id'] != current_user_id and not is_superadmin:
        return {'success': False, 'error': 'Forbidden'}, 403
    file_id = job.get('file_id')
    process = None
    with _lock:
        state = _active_jobs.get(job_id)
        if state:
            state['stopped'] = True
            state['status'] = 'STOPPED'
            state['running'] = 0
            process = state.get('process')
    pid = job.get('process_id')
    if pid and psutil.pid_exists(pid):
        _kill_process_tree(pid, force=True)
    if process:
        _kill_process_tree(process)
        try:
            process.wait(timeout=3)
        except subprocess.TimeoutExpired:
            _kill_process_tree(process, force=True)
            try:
                process.wait(timeout=2)
            except Exception:
                pass
    finalize_job(job_id, status='STOPPED', error_message='Scraper stopped by user.')
    if file_id:
        files_repo.set_working(file_id, 0)
    return {'success': True, 'job_id': job_id, 'status': 'STOPPED', 'message': 'Scraper stopped successfully.'}, 200


def stop_file(file_id, current_user_id, is_superadmin=False):
    active = get_active_log(file_id)
    if active and active.get('job_id'):
        return stop_job(active['job_id'], current_user_id, is_superadmin=is_superadmin)
    files_repo.set_working(file_id, 0)
    return {'success': True, 'file_id': file_id, 'message': 'Scraper is not running.'}, 200


def _watchdog_loop():
    while True:
        try:
            time.sleep(15)
            now = datetime.now(timezone.utc)
            with _lock:
                managed_jobs = list(_active_jobs.items())
            for jid, state in managed_jobs:
                proc = state.get('process')
                pid = state.get('pid')
                started_at = state.get('started_at')
                if started_at and (now - started_at).total_seconds() > MAX_RUNTIME_SECONDS:
                    logger.warning('Job %s exceeded 6 hours. Auto-terminating...', jid)
                    if pid and psutil.pid_exists(pid):
                        _kill_process_tree(pid, force=True)
                    state['timeout_stopped'] = True
                    finalize_job(jid, status='STOPPED',
                                 error_message='Automatically stopped after maximum runtime of 6 hours.')
                    continue
                if proc and proc.poll() is not None and state.get('status') == 'RUNNING':
                    ret = proc.poll()
                    final_st = 'SUCCESS' if ret == 0 else 'FAILED'
                    err = None if ret == 0 else f'Process exited with return code {ret}.'
                    finalize_job(jid, status=final_st, error_message=err)
            conn = get_connection()
            try:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        UPDATE logTbl SET status='STOPPED', end_time=NOW(),
                        error_message='Automatically stopped after maximum runtime of 6 hours.'
                        WHERE status='RUNNING' AND end_time IS NULL
                        AND start_time < NOW() - INTERVAL 6 HOUR
                    """)
            finally:
                conn.close()
        except Exception:
            pass


_watchdog_thread_started = False


def start_watchdog():
    global _watchdog_thread_started
    if not _watchdog_thread_started:
        _watchdog_thread_started = True
        threading.Thread(target=_watchdog_loop, daemon=True).start()


start_watchdog()
