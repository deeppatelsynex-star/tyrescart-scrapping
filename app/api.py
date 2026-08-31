"""
app/api.py - Unified REST/JSON API layer.

Combines all three API surfaces of the application into one file:
  - /tcsadmin/api/*    scraper admin backend (register_tcsadmin_api_routes)
  - /visionadmin/api/* CMS backend (register_visionadmin_api_routes)
  - /api/*             public client-facing API (register_client_api_routes)

register_api_routes(app) registers all three. Page-rendering routes for
each of the three areas live in their own dedicated files instead:
  - app/scraperapp/tcsadmin.py       (tcsadmin pages)
  - app/visionadmin/Visionadminroute.py (visionadmin pages)
  - app/siteapp/clientroute.py       (client pages, site_bp blueprint)
"""

import csv
from datetime import datetime
import io
import json
from collections import OrderedDict
import math
import os
import queue
import re
import secrets
import subprocess
import sys
import threading
import time
import uuid
import zipfile

from flask import Response, jsonify, render_template, request, send_file, send_from_directory, session, stream_with_context
from openpyxl import Workbook, load_workbook
import pymysql
from werkzeug.utils import secure_filename

from auth import (
    VALID_ROLES,
    bit_to_bool,
    get_user_by_id,
    has_superadmin,
    hash_password,
    list_active_users,
    list_deleted_users,
    login_required_api,
    require_csrf,
    role_required_api,
    serialize_user,
    to_ist_12h,
    verify_password,
)
from db import get_connection
from models.blog import Blog
from models.page import Page
from models.page_section import PageSection
from siteapp.clientroute import _get_locale

EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
# This file is app/api.py, so the project root (where scrapers/ and tmp/
# live) is one directory up.
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMP_SCRAPERS_DIR = os.path.join(BASE_DIR, 'tmp', 'scrapers')
sys.path.insert(0, os.path.join(BASE_DIR, 'scrapers'))

import file_scraper_runner
import files_repo
import job_manager
import reports_repo
from scraper_config import SCRIPT_MAP
from scraper_input import (
    InputError,
    build_entries,
    extract_input_source,
    format_invalid_url_message,
    format_unsupported_message,
    parse_csv_urls,
    parse_text_urls,
    validate_url_list,
)
from scraper_status_utils import build_status_summary, parse_status_line


class ScraperSession:
    """In-memory state for one browser session's ad-hoc scraping runs."""
    def __init__(self):
        self.lock = threading.Lock()
        self.process = None
        self.url_statuses = []
        self.job_status = 'idle'  # idle | running | completed_unseen | failed_unseen
        self.stopped = False
        self.job_id = None
        self.output_file = None
        self.thread = None
        self.pending_groups = []
        self.skipped = {'invalid': [], 'unsupported': []}


_scraper_sessions = {}
_sessions_lock = threading.Lock()


def get_scraper_session():
    sid = session.get('sid')
    if not sid:
        sid = secrets.token_hex(16)
        session['sid'] = sid
    with _sessions_lock:
        if sid not in _scraper_sessions:
            _scraper_sessions[sid] = ScraperSession()
        return _scraper_sessions[sid]


def get_xlsx_info(output_file):
    if not output_file or not os.path.exists(output_file):
        return 0, set()
    try:
        wb = load_workbook(output_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return 0, set()
        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        url_col_idx = None
        for candidate in ('url', 'product url', 'link', 'product link'):
            if candidate in header:
                url_col_idx = header.index(candidate)
                break
        urls = set()
        count = 0
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                count += 1
                if url_col_idx is not None and url_col_idx < len(row):
                    val = row[url_col_idx]
                    if val:
                        urls.add(str(val).strip())
        return count, urls
    except Exception:
        return 0, set()


def _record_status_line(state, cleaned_line):
    parsed_status = parse_status_line(cleaned_line)
    if not parsed_status:
        return
    with state.lock:
        existing = next((item for item in state.url_statuses if item['url'] == parsed_status['url']), None)
        if existing:
            existing['status'] = parsed_status['status']
            if parsed_status.get('parent'):
                existing['parent'] = parsed_status['parent']
            if parsed_status.get('type'):
                existing['type'] = parsed_status['type']
        else:
            state.url_statuses.append({
                'url': parsed_status['url'],
                'status': parsed_status['status'],
                'parent': parsed_status.get('parent') or '',
                'type': parsed_status.get('type') or 'root',
            })


def _merge_xlsx_outputs(source_paths, destination):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    header_written = False

    for path in source_paths:
        try:
            src_wb = load_workbook(path, read_only=True)
        except Exception:
            continue
        try:
            rows = src_wb.active.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            if not header_written:
                ws.append(list(header))
                header_written = True
            for row in rows:
                ws.append(list(row))
        finally:
            src_wb.close()

    if not header_written:
        ws.append(['No data'])

    os.makedirs(os.path.dirname(destination), exist_ok=True)
    wb.save(destination)


def _run_job_groups(state):
    group_outputs = []
    had_failure = False

    for group in state.pending_groups:
        with state.lock:
            if state.stopped:
                break

        process = subprocess.Popen(
            [sys.executable, '-u', group['script_path'], group['output_path'], group['input_path']],
            cwd=BASE_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        with state.lock:
            state.process = process

        for line in iter(process.stdout.readline, ''):
            if not line:
                break
            _record_status_line(state, line.rstrip('\n'))
        process.stdout.close()
        process.wait()

        with state.lock:
            state.process = None
        if process.returncode not in (0, None) and not state.stopped:
            had_failure = True
        if os.path.exists(group['output_path']):
            group_outputs.append(group['output_path'])

    _merge_xlsx_outputs(group_outputs, state.output_file)

    for group in state.pending_groups:
        for path in (group['input_path'], group['output_path']):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass

    with state.lock:
        if state.stopped:
            state.job_status = 'idle'
        elif had_failure:
            state.job_status = 'failed_unseen'
        else:
            state.job_status = 'completed_unseen'


def _parse_urls_text(urls_text):
    raw_urls = parse_text_urls(urls_text)
    urls, errors = validate_url_list(raw_urls)
    if errors:
        lines = [f"Invalid URL on row {e['row']}: {e['value']}" for e in errors]
        return [], 'Invalid URL(s):\n' + '\n'.join(lines)
    if not urls:
        return [], 'At least one valid URL is required.'
    return urls, None


def _format_scraper_output_filename(site_name, extension='xlsx'):
    clean_site = re.sub(r'[^A-Za-z0-9]+', '_', site_name or '').strip('_').upper() or 'SCRAPER'
    today = datetime.now().strftime('%d-%m-%Y')
    return f"{clean_site}_{today}.{extension}"


def register_tcsadmin_api_routes(app):
    """Registers all REST, JSON, and Scraper execution APIs under /tcsadmin/api (and bare /api aliases)."""

    # ==========================================================================
    # ==========================================================================
    # 1. User, Profile & Authentication APIs (admin_users table)
    # ==========================================================================

    @app.route('/visionadmin/api/me')
    @app.route('/tcsadmin/api/me')
    @app.route('/api/me')
    def api_me():
        user_id = session.get('admin_user_id') or session.get('user_id') or session.get('userid') or session.get('id')
        email = session.get('email') or session.get('Email')
        if not user_id and not email:
            return jsonify({'error': 'Authentication required.'}), 401

        from visionadmin.admin_auth import get_admin_user_by_id, get_admin_user_by_email, serialize_admin_user
        admin_u = None
        if user_id:
            admin_u = get_admin_user_by_id(user_id)
        if not admin_u and email:
            admin_u = get_admin_user_by_email(email)
            if admin_u:
                session['admin_user_id'] = admin_u['id']
                session['user_id'] = admin_u['id']
                session['userid'] = admin_u['id']
                session['id'] = admin_u['id']
                session['name'] = admin_u['name']
                session['Name'] = admin_u['name']
                session['email'] = admin_u['email']
                session['Email'] = admin_u['email']
                session['role'] = 'SuperAdmin' if admin_u['role'] in ('super_admin', 'superadmin', 'SuperAdmin') else ('Admin' if admin_u['role'] in ('manager', 'admin', 'Admin') else 'User')
                session['admin_role'] = admin_u['role']
                session['is_visionadmin'] = True
                session['logged_in'] = True

        if not admin_u:
            return jsonify({'error': 'Authentication required.'}), 401

        role_disp = 'SuperAdmin' if admin_u.get('role') in ('super_admin', 'superadmin', 'SuperAdmin') else ('Admin' if admin_u.get('role') in ('manager', 'admin', 'Admin') else 'User')

        return jsonify({
            'user': {
                'userid': admin_u['id'],
                'id': admin_u['id'],
                'name': admin_u['name'],
                'Name': admin_u['name'],
                'email': admin_u['email'],
                'Email': admin_u['email'],
                'role': role_disp,
                'Role': role_disp,
                'admin_role': admin_u.get('role'),
                'status': bool(admin_u.get('is_active', 1)),
                'avatar': None,
                'createdAt': str(admin_u.get('created_at', '')),
                'updatedAt': str(admin_u.get('updated_at', '')),
            },
            'csrfToken': session.get('csrf_token')
        })

    @app.route('/visionadmin/api/profile', methods=['PUT'])
    @app.route('/tcsadmin/api/profile', methods=['PUT'])
    @app.route('/api/profile', methods=['PUT'])
    @require_csrf
    def api_update_profile():
        user_id = session.get('admin_user_id') or session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Authentication required.'}), 401

        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip().lower()

        if not name:
            return jsonify({'error': 'Name is required.'}), 400
        if not email or not EMAIL_RE.match(email):
            return jsonify({'error': 'A valid email is required.'}), 400

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                try:
                    cursor.execute(
                        'UPDATE `admin_users` SET `name` = %s, `email` = %s, `updated_at` = NOW() WHERE `id` = %s',
                        (name, email, user_id),
                    )
                    conn.commit()
                except pymysql.err.IntegrityError:
                    return jsonify({'error': 'That email is already in use.'}), 409
        finally:
            conn.close()

        session['name'] = name
        session['email'] = email

        return jsonify({
            'user': {
                'id': user_id,
                'userid': user_id,
                'name': name,
                'Name': name,
                'email': email,
                'Email': email,
                'role': session.get('role'),
                'Role': session.get('role')
            }
        })

    @app.route('/visionadmin/api/profile/avatar', methods=['DELETE'])
    @app.route('/tcsadmin/api/profile/avatar', methods=['DELETE'])
    @app.route('/api/profile/avatar', methods=['DELETE'])
    @require_csrf
    def api_delete_avatar():
        return jsonify({'success': True})

    @app.route('/visionadmin/api/change-password', methods=['POST'])
    @app.route('/visonadmin/api/change-password', methods=['POST'])
    @app.route('/tcsadmin/api/change-password', methods=['POST'])
    @app.route('/api/change-password', methods=['POST'])
    @require_csrf
    def api_change_password():
        user_id = session.get('admin_user_id') or session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Authentication required.'}), 401

        fail_count = session.get('pwd_fail_count', 0)
        if fail_count >= 5:
            return jsonify({'error': 'Too many failed attempts. Please try again later.'}), 429

        data = request.get_json(silent=True) or {}
        current_password = data.get('current_password') or ''
        new_password = data.get('new_password') or ''
        confirm_password = data.get('confirm_password') or ''

        if not current_password or not new_password or not confirm_password:
            return jsonify({'error': 'All fields are required.'}), 400
        if new_password != confirm_password:
            return jsonify({'error': 'New password and confirmation do not match.'}), 400
        if len(new_password) < 8:
            return jsonify({'error': 'New password must be at least 8 characters.'}), 400

        from visionadmin.admin_auth import get_admin_user_by_id, verify_admin_password, update_admin_user_password
        admin_u = get_admin_user_by_id(user_id)
        if not admin_u or not verify_admin_password(current_password, admin_u.get('password', '')):
            session['pwd_fail_count'] = fail_count + 1
            return jsonify({'error': 'Current password is incorrect.'}), 400

        update_admin_user_password(user_id, new_password)
        session['pwd_fail_count'] = 0
        return jsonify({'message': 'Password updated successfully.'})

    @app.route('/tcsadmin/api/profile/delete-account', methods=['POST', 'DELETE'])
    @app.route('/api/profile/delete-account', methods=['POST', 'DELETE'])
    @app.route('/visionadmin/api/profile/delete-account', methods=['POST', 'DELETE'])
    @app.route('/api/profile/delete', methods=['POST', 'DELETE'])
    @login_required_api
    @require_csrf
    def api_self_delete_account():
        user_id = session.get('admin_user_id') or session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Authentication required.'}), 401

        from visionadmin.admin_auth import toggle_admin_user_status, count_super_admins, get_admin_user_by_id
        target = get_admin_user_by_id(user_id)
        if not target:
            session.clear()
            return jsonify({'error': 'User not found.'}), 404

        if target.get('role') == 'super_admin' and count_super_admins() <= 1:
            return jsonify({'error': 'Cannot delete the only remaining Super Administrator.'}), 400

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute('UPDATE `admin_users` SET `is_active` = 0, `updated_at` = NOW() WHERE `id` = %s', (user_id,))
                conn.commit()
        finally:
            conn.close()

        session.clear()
        return jsonify({
            'success': True,
            'message': 'Your account has been deactivated successfully.',
            'redirect': '/visionadmin/login'
        })

    # ==========================================================================
    # 2. Admin User Management APIs (Alias to admin_users table)
    # ==========================================================================

    @app.route('/tcsadmin/api/admin/users', methods=['GET'])
    @app.route('/api/admin/users', methods=['GET'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin', 'super_admin', 'manager')
    def api_admin_list_users():
        return jsonify({'users': [serialize_user(u) for u in list_active_users()]})

    @app.route('/tcsadmin/api/admin/users/trash', methods=['GET'])
    @app.route('/api/admin/users/trash', methods=['GET'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin', 'super_admin', 'manager')
    def api_admin_list_trash():
        return jsonify({'users': [serialize_user(u) for u in list_deleted_users()]})

    @app.route('/tcsadmin/api/admin/users', methods=['POST'])
    @app.route('/api/admin/users', methods=['POST'])
    @login_required_api
    @role_required_api('SuperAdmin')
    @require_csrf
    def api_admin_create_user():
        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        password = data.get('password') or ''
        role = (data.get('role') or '').strip()
        status = data.get('status', True)

        if not name:
            return jsonify({'error': 'Name is required.'}), 400
        if not email or not EMAIL_RE.match(email):
            return jsonify({'error': 'A valid email is required.'}), 400
        if not password or len(password) < 8:
            return jsonify({'error': 'Password must be at least 8 characters.'}), 400
        if role not in VALID_ROLES:
            return jsonify({'error': f"Role must be one of: {', '.join(VALID_ROLES)}."}), 400
        if role == 'SuperAdmin' and has_superadmin():
            return jsonify({'error': 'A SuperAdmin already exists. Only one SuperAdmin account is allowed.'}), 409

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                try:
                    cursor.execute(
                        'INSERT INTO userTbl (Name, Email, password, Status, IsDeleted, Role) '
                        'VALUES (%s, %s, %s, %s, 0, %s)',
                        (name, email, hash_password(password), 1 if status else 0, role),
                    )
                except pymysql.err.IntegrityError:
                    return jsonify({'error': 'A user with that email already exists.'}), 409
                new_user_id = cursor.lastrowid
        finally:
            conn.close()

        return jsonify({'user': serialize_user(get_user_by_id(new_user_id))}), 201

    @app.route('/tcsadmin/api/admin/users/<int:user_id>', methods=['PUT'])
    @app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin')
    @require_csrf
    def api_admin_update_user(user_id):
        target = get_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'User not found.'}), 404

        actor_role = session.get('role')
        if target['Role'] == 'SuperAdmin' and actor_role != 'SuperAdmin':
            return jsonify({'error': 'Only a SuperAdmin can modify a SuperAdmin account.'}), 403

        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        role = (data.get('role') or '').strip()
        password = data.get('password') or ''
        status = data.get('status', True)

        if not name:
            return jsonify({'error': 'Name is required.'}), 400
        if not email or not EMAIL_RE.match(email):
            return jsonify({'error': 'A valid email is required.'}), 400
        if role not in VALID_ROLES:
            return jsonify({'error': f"Role must be one of: {', '.join(VALID_ROLES)}."}), 400
        if role == 'SuperAdmin' and actor_role != 'SuperAdmin':
            return jsonify({'error': 'Only a SuperAdmin can grant the SuperAdmin role.'}), 403
        if role == 'SuperAdmin' and target['Role'] != 'SuperAdmin' and has_superadmin():
            return jsonify({'error': 'A SuperAdmin already exists. Only one SuperAdmin account is allowed.'}), 409
        if password and len(password) < 8:
            return jsonify({'error': 'Password must be at least 8 characters.'}), 400

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                try:
                    if password:
                        cursor.execute(
                            'UPDATE userTbl SET Name = %s, Email = %s, Role = %s, Status = %s, password = %s '
                            'WHERE userid = %s',
                            (name, email, role, 1 if status else 0, hash_password(password), user_id),
                        )
                    else:
                        cursor.execute(
                            'UPDATE userTbl SET Name = %s, Email = %s, Role = %s, Status = %s WHERE userid = %s',
                            (name, email, role, 1 if status else 0, user_id),
                        )
                except pymysql.err.IntegrityError:
                    return jsonify({'error': 'That email is already in use.'}), 409
        finally:
            conn.close()

        updated = get_user_by_id(user_id)
        if user_id == session.get('user_id'):
            session['name'] = updated['Name']
            session['email'] = updated['Email']
            session['role'] = updated['Role']

        return jsonify({'user': serialize_user(updated)})

    @app.route('/tcsadmin/api/admin/users/<int:user_id>', methods=['DELETE'])
    @app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin')
    @require_csrf
    def api_admin_delete_user(user_id):
        target = get_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'User not found.'}), 404

        if target['Role'] == 'SuperAdmin' and user_id != session.get('user_id'):
            return jsonify({'error': 'SuperAdmin accounts can only be deleted by the account owner.'}), 403

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    'UPDATE userTbl SET IsDeleted = 1, deleted_at = UTC_TIMESTAMP() WHERE userid = %s',
                    (user_id,),
                )
        finally:
            conn.close()

        is_self = (user_id == session.get('user_id'))
        if is_self:
            session.clear()
            return jsonify({'message': 'Your account has been deleted.', 'selfDeleted': True, 'redirect': '/tcsadmin/login'})

        return jsonify({'message': 'User deleted.'})

    @app.route('/tcsadmin/api/admin/users/<int:user_id>/recover', methods=['POST'])
    @app.route('/api/admin/users/<int:user_id>/recover', methods=['POST'])
    @login_required_api
    @role_required_api('SuperAdmin')
    @require_csrf
    def api_admin_recover_user(user_id):
        target = get_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'User not found.'}), 404
        if not bit_to_bool(target['IsDeleted']):
            return jsonify({'error': 'This account is not deleted.'}), 400

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute('UPDATE userTbl SET IsDeleted = 0, deleted_at = NULL WHERE userid = %s', (user_id,))
        finally:
            conn.close()

        return jsonify({'message': 'User recovered.'})

    # ==========================================================================
    # 3. File / Registered Scraper APIs
    # ==========================================================================

    @app.route('/tcsadmin/api/files/running')
    @app.route('/api/files/running')
    @login_required_api
    def api_list_running_files():
        rows, _ = files_repo.list_files(per_page=200)
        running = [
            {'fileId': r['file_id'], 'siteName': r['site_name']}
            for r in rows
            if file_scraper_runner.is_running(r['file_id'])
        ]
        return jsonify({'files': running})

    @app.route('/tcsadmin/api/files')
    @app.route('/api/files')
    @login_required_api
    def api_list_files():
        search = request.args.get('search', '').strip() or None
        raw_trash = request.args.get('trash')
        is_deleted = None
        if raw_trash is not None:
            is_deleted = raw_trash.lower() in ('1', 'true', 'yes')

        try:
            page = int(request.args.get('page', 1))
        except ValueError:
            page = 1
        try:
            per_page = int(request.args.get('perPage', 20))
        except ValueError:
            per_page = 20

        rows, total = files_repo.list_files(search=search, is_deleted=is_deleted, page=page, per_page=per_page)
        serialized = []
        user_id = session.get('user_id')
        active_map = job_manager.get_all_active_jobs_map()
        all_outputs = file_scraper_runner.get_all_output_paths()

        for r in rows:
            fid = r['file_id']
            item = files_repo.serialize_file(r)
            active = active_map.get(fid)
            if active:
                item['working'] = True
                item['is_owner'] = (active.get('user_id') == user_id)
            else:
                item['working'] = False
                item['is_owner'] = True
            item['outputAvailable'] = bool(all_outputs.get(fid))
            serialized.append(item)

        any_running = bool(active_map)
        has_any_output = bool(all_outputs)

        return jsonify({
            'files': serialized,
            'total': total,
            'page': page,
            'perPage': per_page,
            'anyRunning': any_running,
            'hasAnyOutput': has_any_output,
        })

    @app.route('/tcsadmin/api/files', methods=['POST'])
    @app.route('/api/files', methods=['POST'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin')
    @require_csrf
    def api_create_file():
        data = request.get_json(silent=True) or {}
        site_name = (data.get('siteName') or '').strip()
        python_file_path = (data.get('pythonFilePath') or '').strip()
        logo = (data.get('logo') or '').strip() or None
        urls_text = data.get('urlsText') or ''

        if not site_name:
            return jsonify({'error': 'Name is required.'}), 400
        if not python_file_path:
            return jsonify({'error': 'Python file is required.'}), 400

        urls, url_error = _parse_urls_text(urls_text)
        if url_error:
            return jsonify({'error': url_error}), 400

        try:
            created_by = session.get('user_id')
            file_id = files_repo.create_file(logo, site_name, python_file_path, created_by=created_by)
        except files_repo.FileValidationError as exc:
            return jsonify({'error': str(exc)}), 400

        files_repo.set_urls(file_id, urls)
        return jsonify({'file': files_repo.serialize_file(files_repo.get_file(file_id))}), 201

    @app.route('/tcsadmin/api/files/<int:file_id>', methods=['PUT'])
    @app.route('/api/files/<int:file_id>', methods=['PUT'])
    @login_required_api
    @require_csrf
    def api_update_file(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404
        if file_scraper_runner.is_running(file_id):
            return jsonify({'error': 'Stop this scraper before editing it.'}), 409

        data = request.get_json(silent=True) or {}
        site_name = (data.get('siteName') or '').strip()
        python_file_path = (data.get('pythonFilePath') or '').strip()
        logo = (data.get('logo') or '').strip() or None
        urls_text = data.get('urlsText') or ''

        if not site_name:
            return jsonify({'error': 'Name is required.'}), 400
        if not python_file_path:
            return jsonify({'error': 'Python file is required.'}), 400

        urls, url_error = _parse_urls_text(urls_text)
        if url_error:
            return jsonify({'error': url_error}), 400

        try:
            files_repo.update_file(file_id, logo, site_name, python_file_path)
        except files_repo.FileValidationError as exc:
            return jsonify({'error': str(exc)}), 400

        files_repo.set_urls(file_id, urls)
        return jsonify({'file': files_repo.serialize_file(files_repo.get_file(file_id))})

    @app.route('/tcsadmin/api/files/<int:file_id>', methods=['DELETE'])
    @app.route('/api/files/<int:file_id>', methods=['DELETE'])
    @login_required_api
    @require_csrf
    def api_delete_file(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404
        if file_scraper_runner.is_running(file_id):
            return jsonify({'error': 'Stop this scraper before deleting it.'}), 409

        files_repo.delete_file(file_id)
        return jsonify({'message': 'Scraper permanently deleted.'})

    @app.route('/tcsadmin/api/files/<int:file_id>/toggle-status', methods=['POST'])
    @app.route('/api/files/<int:file_id>/toggle-status', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_toggle_file_status(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404

        if file_scraper_runner.is_running(file_id):
            return jsonify({'error': 'Stop this scraper before disabling it.'}), 409

        data = request.get_json(silent=True) or {}
        currently_deleted = files_repo.bit_to_bool(record.get('is_deleted'))

        if 'enabled' in data:
            new_enabled = bool(data['enabled'])
        else:
            new_enabled = currently_deleted

        files_repo.set_file_enabled(file_id, new_enabled)
        updated = files_repo.get_file(file_id)
        return jsonify({
            'success': True,
            'isEnabled': new_enabled,
            'isDeleted': not new_enabled,
            'message': 'Scraper enabled.' if new_enabled else 'Scraper disabled.',
            'file': files_repo.serialize_file(updated),
        })

    @app.route('/tcsadmin/api/files/<int:file_id>/restore', methods=['POST'])
    @app.route('/api/files/<int:file_id>/restore', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_restore_file(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404

        files_repo.restore_file(file_id)
        updated = files_repo.get_file(file_id)
        return jsonify({
            'success': True,
            'message': 'Scraper restored to Active list.',
            'file': files_repo.serialize_file(updated),
        })

    @app.route('/tcsadmin/api/files/upload-script', methods=['POST'])
    @app.route('/api/files/upload-script', methods=['POST'])
    @login_required_api
    @role_required_api('SuperAdmin', 'Admin')
    @require_csrf
    def api_upload_file_script():
        if not request.files or 'file' not in request.files:
            return jsonify({'error': 'No file uploaded.'}), 400

        upload = request.files['file']
        candidate_name = (upload.filename or '').strip()
        if candidate_name:
            existing = files_repo.get_file_by_path(candidate_name)
            if existing and file_scraper_runner.is_running(existing['file_id']):
                return jsonify({'error': 'Stop this scraper before replacing its Python file.'}), 409

        try:
            filename = files_repo.save_uploaded_script(upload)
        except files_repo.FileValidationError as exc:
            return jsonify({'error': str(exc)}), 400

        return jsonify({'fileName': filename})

    @app.route('/tcsadmin/api/files/parse-urls', methods=['POST'])
    @app.route('/api/files/parse-urls', methods=['POST'])
    @login_required_api
    def api_parse_urls():
        if request.files and 'file' in request.files:
            upload = request.files['file']
            filename = (upload.filename or '').lower()
            if not filename.endswith('.csv'):
                return jsonify({'error': 'Please upload a .csv file.'}), 400
            raw_bytes = upload.read()
            if not raw_bytes:
                return jsonify({'error': 'The uploaded file is empty.'}), 400
            raw_urls = [url for url, _declared_type in parse_csv_urls(raw_bytes)]
        else:
            data = request.get_json(silent=True) or {}
            raw_urls = parse_text_urls(data.get('text') or '')

        urls, errors = validate_url_list(raw_urls)
        if not urls:
            return jsonify({'error': 'No valid URLs were found.', 'errors': errors}), 400

        return jsonify({'urls': urls, 'errors': errors})

    # ==========================================================================
    # 4. Scraper Execution, Job Manager & Streaming APIs
    # ==========================================================================

    @app.route('/tcsadmin/api/scraper/start', methods=['POST'])
    @app.route('/api/scraper/start', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_scraper_job_start():
        data = request.get_json(silent=True) or {}
        file_id = data.get('file_id') or request.form.get('file_id') or request.args.get('file_id')
        if not file_id:
            return jsonify({'success': False, 'error': 'Missing file_id parameter.'}), 400
        try:
            file_id = int(file_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid file_id.'}), 400

        user_id = session.get('user_id')
        result = job_manager.start_job(file_id, user_id=user_id)
        if not result.get('success'):
            return jsonify(result), 409
        return jsonify(result)

    @app.route('/tcsadmin/api/scraper/file/<int:file_id>/active-job')
    @app.route('/api/scraper/file/<int:file_id>/active-job')
    @login_required_api
    def api_scraper_file_active_job(file_id):
        user_id = session.get('user_id')
        active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)
        return jsonify(active_info)

    @app.route('/tcsadmin/api/scraper/job/<string:job_id>/status')
    @app.route('/api/scraper/job/<string:job_id>/status')
    @login_required_api
    def api_scraper_job_status(job_id):
        user_id = session.get('user_id')
        data, code = job_manager.get_job_status(job_id, current_user_id=user_id)
        return jsonify(data), code

    @app.route('/tcsadmin/api/scraper/job/<string:job_id>/urls')
    @app.route('/api/scraper/job/<string:job_id>/urls')
    @login_required_api
    def api_scraper_job_urls(job_id):
        user_id = session.get('user_id')
        urls, code = job_manager.get_job_urls(job_id, current_user_id=user_id)
        if code != 200:
            return jsonify(urls), code
        summary = build_status_summary(urls)
        return jsonify({
            'job_id': job_id,
            'statuses': urls,
            'summary': summary,
            'count': len(urls),
        })

    @app.route('/tcsadmin/api/scraper/job/<string:job_id>/events')
    @app.route('/api/scraper/job/<string:job_id>/events')
    @login_required_api
    def api_scraper_job_events(job_id):
        user_id = session.get('user_id')
        job = job_manager.get_log_by_job_id(job_id)
        if not job:
            with job_manager._lock:
                state = job_manager._active_jobs.get(job_id)
            if not state:
                return jsonify({'error': 'Job not found.'}), 404
            if state['started_by_user_id'] != user_id and session.get('role') != 'SuperAdmin':
                return jsonify({'error': 'Forbidden'}), 403
        elif job['user_id'] != user_id and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden'}), 403

        def event_stream():
            q = job_manager.subscribe_sse(job_id)
            try:
                status_data, _ = job_manager.get_job_status(job_id, current_user_id=user_id)
                urls_data, _ = job_manager.get_job_urls(job_id, current_user_id=user_id)
                initial_payload = {
                    'type': 'snapshot',
                    'summary': status_data,
                    'statuses': urls_data
                }
                yield f"data: {json.dumps(initial_payload)}\n\n"

                while True:
                    try:
                        event = q.get(timeout=15.0)
                        yield f"data: {json.dumps(event)}\n\n"
                        if event.get('done') or (event.get('type') == 'status' and event.get('status') in ('SUCCESS', 'STOPPED', 'FAILED', 'FAIL')):
                            break
                    except queue.Empty:
                        yield ": ping\n\n"
            except GeneratorExit:
                pass
            finally:
                job_manager.unsubscribe_sse(job_id, q)

        return Response(
            stream_with_context(event_stream()),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no',
                'Connection': 'keep-alive',
            }
        )

    @app.route('/tcsadmin/api/scraper/job/<string:job_id>/stop', methods=['POST'])
    @app.route('/api/scraper/job/<string:job_id>/stop', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_scraper_job_stop(job_id):
        user_id = session.get('user_id')
        is_superadmin = (session.get('role') == 'SuperAdmin')
        result, code = job_manager.stop_job(job_id, current_user_id=user_id, is_superadmin=is_superadmin)
        return jsonify(result), code

    @app.route('/tcsadmin/api/files/<int:file_id>/start', methods=['POST'])
    @app.route('/api/files/<int:file_id>/start', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_start_file(file_id):
        user_id = session.get('user_id')
        result = job_manager.start_job(file_id, user_id=user_id)
        if not result.get('success'):
            return jsonify(result), 409
        return jsonify(result)

    @app.route('/tcsadmin/api/files/<int:file_id>/stop', methods=['POST'])
    @app.route('/api/files/<int:file_id>/stop', methods=['POST'])
    @login_required_api
    @require_csrf
    def api_stop_file(file_id):
        user_id = session.get('user_id')
        is_superadmin = (session.get('role') == 'SuperAdmin')
        result, code = job_manager.stop_file(file_id, current_user_id=user_id, is_superadmin=is_superadmin)
        return jsonify(result), code

    @app.route('/tcsadmin/api/files/<int:file_id>/status')
    @app.route('/api/files/<int:file_id>/status')
    @login_required_api
    def api_file_status(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404

        user_id = session.get('user_id')
        active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)

        if active_info.get('job_id'):
            job_status, code = job_manager.get_job_status(active_info['job_id'], current_user_id=user_id)
            if code == 200:
                job_status['is_owner'] = True
                job_status['working'] = bool(active_info.get('has_active_job'))
                job_status['siteName'] = record['site_name']
                job_status['fileId'] = file_id
                return jsonify(job_status)

        output_path = file_scraper_runner.get_output_path(file_id)
        return jsonify({
            'job_id': None,
            'running': False,
            'working': False,
            'done': True,
            'is_owner': True,
            'siteName': record['site_name'],
            'fileId': file_id,
            'outputAvailable': bool(output_path and os.path.exists(output_path)),
            'total_product_urls': 0,
            'written_to_xlsx': 0,
            'pending': 0,
            'running_count': 0,
            'blocked': 0,
            'main_url_done': 0,
            'product_url_done': 0,
            'progress_percent': 0.0,
        })

    @app.route('/tcsadmin/api/files/<int:file_id>/url-statuses')
    @app.route('/api/files/<int:file_id>/url-statuses')
    @login_required_api
    def api_file_url_statuses(file_id):
        user_id = session.get('user_id')
        active_info = job_manager.get_active_job_for_file(file_id, current_user_id=user_id)

        if active_info['has_active_job'] and not active_info['is_owner']:
            return jsonify({'statuses': [], 'summary': {}, 'xlsx_count': 0, 'error': 'Forbidden'}), 403

        if active_info.get('job_id') and active_info.get('is_owner', True):
            urls, code = job_manager.get_job_urls(active_info['job_id'], current_user_id=user_id)
            if code == 200 and urls:
                summary = build_status_summary(urls)
                return jsonify({
                    'statuses': urls,
                    'summary': summary,
                    'xlsx_count': summary.get('written_to_xlsx', 0),
                })

        urls = file_scraper_runner.get_statuses(file_id)
        summary = build_status_summary(urls)
        return jsonify({
            'statuses': urls,
            'summary': summary,
            'xlsx_count': summary.get('written_to_xlsx', 0),
        })

    @app.route('/tcsadmin/api/files/<int:file_id>/download')
    @app.route('/api/files/<int:file_id>/download')
    @login_required_api
    def api_file_download(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404

        output_path = file_scraper_runner.get_output_path(file_id)
        if not output_path or not os.path.exists(output_path):
            return jsonify({'error': 'No output available for this scraper yet. Run it first.'}), 404

        filename = _format_scraper_output_filename(record['site_name'])
        return send_file(
            output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    @app.route('/tcsadmin/api/files/download-zip')
    @app.route('/api/files/download-zip')
    @login_required_api
    def api_files_download_zip():
        if file_scraper_runner.running_count() > 0:
            return jsonify({
                'error': 'Scraping is in progress. Please wait until all scrapers finish before downloading ZIP.',
                'anyRunning': True,
            }), 409

        all_outputs = file_scraper_runner.get_all_output_paths()
        if not all_outputs:
            return jsonify({'error': 'No completed scraper reports available to download.'}), 404

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            used_names = set()
            for file_id, file_path in all_outputs.items():
                if not os.path.exists(file_path):
                    continue
                record = files_repo.get_file(file_id)
                site_name = record['site_name'] if record else f"scraper_{file_id}"
                base_filename = _format_scraper_output_filename(site_name)
                filename = base_filename
                counter = 1
                while filename in used_names:
                    root, ext = os.path.splitext(base_filename)
                    filename = f"{root}_{counter}{ext}"
                    counter += 1
                used_names.add(filename)
                zf.write(file_path, arcname=filename)

        if not used_names:
            return jsonify({'error': 'No valid output files found to package.'}), 404

        zip_buffer.seek(0)
        today = datetime.now().strftime('%d-%m-%Y')
        zip_filename = f"scrapers_output_{today}.zip"
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename,
        )

    @app.route('/tcsadmin/api/files/<int:file_id>/logs')
    @app.route('/api/files/<int:file_id>/logs')
    @login_required_api
    def api_file_logs(file_id):
        record = files_repo.get_file(file_id)
        if not record:
            return jsonify({'error': 'Scraper not found.'}), 404
        rows, total = reports_repo.list_logs(file_id=file_id, per_page=100)
        return jsonify({
            'fileId': file_id,
            'siteName': record.get('site_name') or 'Scraper',
            'logs': [reports_repo.serialize_log(r) for r in rows],
            'total': total,
        })

    # ==========================================================================
    # 5. Reports & Audit Log APIs
    # ==========================================================================

    @app.route('/tcsadmin/api/reports')
    @app.route('/api/reports')
    @login_required_api
    @role_required_api('SuperAdmin', 'super_admin', 'Admin', 'manager')
    def api_list_reports():
        search = request.args.get('search', '').strip() or None
        status = request.args.get('status', '').strip() or None
        user_id_raw = request.args.get('userId', '').strip()
        user_id = int(user_id_raw) if user_id_raw.isdigit() else None
        file_id_raw = request.args.get('fileId', '').strip()
        file_id = int(file_id_raw) if file_id_raw.isdigit() else None
        try:
            page = int(request.args.get('page', 1))
        except ValueError:
            page = 1
        try:
            per_page = int(request.args.get('perPage', 20))
        except ValueError:
            per_page = 20

        rows, total = reports_repo.list_logs(
            search=search,
            status=status,
            user_id=user_id,
            file_id=file_id,
            page=page,
            per_page=per_page,
        )
        stats = reports_repo.get_logs_summary_stats()
        return jsonify({
            'reports': [reports_repo.serialize_log(r) for r in rows],
            'total': total,
            'page': page,
            'perPage': per_page,
            'stats': stats,
        })

    @app.route('/tcsadmin/api/reports/<int:report_id>/download')
    @app.route('/api/reports/<int:report_id>/download')
    @login_required_api
    @role_required_api('SuperAdmin', 'super_admin', 'Admin', 'manager')
    def api_download_report_output(report_id):
        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT * FROM logTbl WHERE id = %s", (report_id,))
                record = cursor.fetchone()
        finally:
            conn.close()

        if not record or not record.get('output_file_path') or not os.path.exists(record['output_file_path']):
            return jsonify({'error': 'Output file not available for this run.'}), 404

        filename = _format_scraper_output_filename(record['scraper'])
        return send_file(
            record['output_file_path'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    # ==========================================================================
    # 6. Legacy / Ad-hoc Scraper APIs
    # ==========================================================================

    @app.route('/tcsadmin/api/scraper/analyze', methods=['POST'])
    @app.route('/api/scraper/analyze', methods=['POST'])
    @login_required_api
    def analyze_scraper_input():
        try:
            raw_items = extract_input_source(request)
        except InputError as exc:
            return jsonify({'error': str(exc)}), 400

        entries, errors, unsupported = build_entries(raw_items)

        return jsonify({
            'entries': entries,
            'errors': errors,
            'unsupported': unsupported,
            'message': format_invalid_url_message(errors) or format_unsupported_message(unsupported) or None,
        })

    @app.route('/tcsadmin/StartScraper', methods=['POST'])
    @app.route('/StartScraper', methods=['POST'])
    @login_required_api
    def start_scraper():
        state = get_scraper_session()

        try:
            raw_items = extract_input_source(request)
        except InputError as exc:
            return jsonify({'error': str(exc)}), 400

        entries, errors, unsupported = build_entries(raw_items)

        if not entries:
            message = (
                format_invalid_url_message(errors)
                or format_unsupported_message(unsupported)
                or 'No valid URLs were provided.'
            )
            return jsonify({'error': message, 'errors': errors, 'unsupported': unsupported}), 400

        with state.lock:
            process_running = state.process is not None and state.process.poll() is None
            if process_running or state.job_status == 'running':
                return jsonify({'error': 'Scraper is already running.'}), 409

            state.url_statuses.clear()
            state.stopped = False
            state.job_id = uuid.uuid4().hex
            state.job_status = 'running'
            state.skipped = {'invalid': errors, 'unsupported': unsupported}

            job_id_short = state.job_id[:8]

            groups_by_type = OrderedDict()
            for entry in entries:
                groups_by_type.setdefault(entry['type'], []).append(entry['url'])

            os.makedirs(TMP_SCRAPERS_DIR, exist_ok=True)
            pending_groups = []
            for url_type, urls in groups_by_type.items():
                input_path = os.path.join(TMP_SCRAPERS_DIR, f'job_{job_id_short}_{url_type}.csv')
                with open(input_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for url in urls:
                        writer.writerow([url])

                pending_groups.append({
                    'type': url_type,
                    'input_path': input_path,
                    'output_path': os.path.join(TMP_SCRAPERS_DIR, f'job_{job_id_short}_{url_type}_output.xlsx'),
                    'script_path': os.path.join(BASE_DIR, 'scrapers', SCRIPT_MAP[url_type]),
                })
            state.pending_groups = pending_groups

            timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
            state.output_file = os.path.join(BASE_DIR, f'pitstoparabia_data_{job_id_short}_{timestamp}.xlsx')

            state.thread = threading.Thread(target=_run_job_groups, args=(state,), daemon=True)
            state.thread.start()

        return jsonify({
            'message': f'Scraper started. Job ID: {state.job_id}',
            'jobId': state.job_id,
            'groups': [g['type'] for g in pending_groups],
            'skipped': {'invalid': errors, 'unsupported': unsupported},
        })

    @app.route('/tcsadmin/stop-scraper', methods=['POST'])
    @app.route('/stop-scraper', methods=['POST'])
    @login_required_api
    def stop_scraper():
        state = get_scraper_session()
        with state.lock:
            process_running = state.process is not None and state.process.poll() is None
            if not process_running and state.job_status != 'running':
                return jsonify({'stopped': False, 'message': 'No scraper process is running.'})

            state.stopped = True
            if process_running:
                try:
                    state.process.terminate()
                    state.process.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    state.process.kill()
                    state.process.wait()
                finally:
                    state.process = None

        return jsonify({'stopped': True, 'message': 'Scraper has been stopped.'})

    @app.route('/tcsadmin/scraper-url-statuses')
    @app.route('/scraper-url-statuses')
    @login_required_api
    def scraper_url_statuses_endpoint():
        state = get_scraper_session()
        with state.lock:
            statuses = list(state.url_statuses)
            running = state.process is not None and state.process.poll() is None
            output_file = state.output_file
            if not running and state.job_status == 'idle' and state.url_statuses:
                state.url_statuses.clear()

        xlsx_count, xlsx_urls = get_xlsx_info(output_file)
        if xlsx_urls:
            if statuses:
                for item in statuses:
                    u = item.get('url', '').strip()
                    if u in xlsx_urls:
                        item['status'] = 'done'
                        item['written_to_xlsx'] = True
            else:
                statuses = [
                    {'url': u, 'status': 'done', 'parent': '', 'type': 'product', 'written_to_xlsx': True}
                    for u in sorted(xlsx_urls)
                ]

        return jsonify({
            'statuses': statuses,
            'summary': build_status_summary(statuses),
            'xlsx_count': xlsx_count,
        })

    @app.route('/tcsadmin/scraper-status')
    @app.route('/scraper-status')
    @login_required_api
    def scraper_status():
        state = get_scraper_session()
        with state.lock:
            running = state.process is not None and state.process.poll() is None
            if running:
                state.job_status = 'running'

            current = state.job_status
            if current in ('completed_unseen', 'failed_unseen'):
                reported_status = 'completed' if current == 'completed_unseen' else 'failed'
                state.job_status = 'idle'
            elif current == 'running':
                reported_status = 'running'
            else:
                reported_status = 'idle'

            has_active_job = reported_status == 'running'
            job_id = state.job_id if has_active_job else None
            output_file = state.output_file

        output_available = bool(output_file and os.path.exists(output_file))

        return jsonify({
            'running': reported_status == 'running',
            'done': reported_status in ('completed', 'failed'),
            'outputAvailable': output_available,
            'outputFile': os.path.basename(output_file) if output_available else '',
            'status': reported_status,
            'hasActiveJob': has_active_job,
            'jobId': job_id,
        })

    @app.route('/tcsadmin/download-output')
    @app.route('/download-output')
    @login_required_api
    def download_output():
        state = get_scraper_session()
        with state.lock:
            output_file = state.output_file

        if not output_file or not os.path.exists(output_file):
            return Response('No output file found.', status=404, mimetype='text/plain')

        return send_from_directory(BASE_DIR, os.path.basename(output_file), as_attachment=True)


def register_visionadmin_api_routes(app):
    """Registers all /visionadmin/api JSON endpoints (uploads, pages, blogs, sections CRUD)."""

    @app.before_request
    def visionadmin_api_auth_guard():
        """Protects all /visionadmin/api/ endpoints with session auth and RBAC against admin_users."""
        path_lower = request.path.lower()
        if path_lower.startswith('/visionadmin/api/') or path_lower.startswith('/visonadmin/api/'):
            user_id = session.get('admin_user_id') or session.get('user_id')
            if not user_id:
                return jsonify({'error': 'Authentication required. Please sign in to VisionAdmin.'}), 401
            role = session.get('role')
            role_norm = str(role).strip().lower().replace('-', '_').replace(' ', '_') if role else ''
            if role_norm not in {'super_admin', 'superadmin', 'manager', 'support', 'admin'} and role not in ('SuperAdmin', 'Admin'):
                return jsonify({'error': 'Unauthorized. VisionAdmin access requires administrator privileges.'}), 403

    # =========================================================================
    # 1. FILE UPLOAD ENDPOINTS
    # =========================================================================

    @app.route('/visionadmin/api/upload-banner', methods=['POST'])
    def visionadmin_upload_banner():
        file = request.files.get('file') or request.files.get('banner')
        if not file or not file.filename:
            return jsonify({'error': 'No image file provided.'}), 400

        allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp', '.svg', '.gif', '.avif'}
        orig_filename = secure_filename(file.filename)
        _, ext = os.path.splitext(orig_filename)
        ext = ext.lower()
        if ext not in allowed_extensions:
            return jsonify({'error': f'Invalid image format "{ext}". Allowed formats: PNG, JPG, JPEG, WEBP, SVG, GIF, AVIF'}), 400

        upload_folder = os.path.join(app.static_folder, 'uploads', 'pages')
        os.makedirs(upload_folder, exist_ok=True)

        unique_name = f"banner_{int(time.time())}_{uuid.uuid4().hex[:8]}{ext}"
        save_path = os.path.join(upload_folder, unique_name)
        file.save(save_path)

        web_url = f"/static/uploads/pages/{unique_name}"
        return jsonify({
            'success': True,
            'url': web_url,
            'filename': unique_name,
            'message': 'Banner image uploaded successfully.'
        })

    @app.route('/visionadmin/api/upload-blog-image', methods=['POST'])
    def visionadmin_upload_blog_image():
        file = request.files.get('file') or request.files.get('image')
        if not file or not file.filename:
            return jsonify({'error': 'No image file provided.'}), 400

        allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp', '.svg', '.gif', '.avif'}
        orig_filename = secure_filename(file.filename)
        _, ext = os.path.splitext(orig_filename)
        ext = ext.lower()
        if ext not in allowed_extensions:
            return jsonify({'error': f'Invalid image format "{ext}". Allowed formats: PNG, JPG, JPEG, WEBP, SVG, GIF, AVIF'}), 400

        upload_folder = os.path.join(app.static_folder, 'uploads', 'blogs')
        os.makedirs(upload_folder, exist_ok=True)

        unique_name = f"blog_{int(time.time())}_{uuid.uuid4().hex[:8]}{ext}"
        save_path = os.path.join(upload_folder, unique_name)
        file.save(save_path)

        web_url = f"/static/uploads/blogs/{unique_name}"
        return jsonify({
            'success': True,
            'url': web_url,
            'filename': unique_name,
            'message': 'Featured blog image uploaded successfully.'
        })

    # =========================================================================
    # 2. PAGES JSON API ENDPOINTS (/visionadmin/api/pages & /visionadmin/api/v1/pages)
    # =========================================================================

    @app.route('/visionadmin/api/pages', methods=['GET'])
    @app.route('/visionadmin/api/v1/pages', methods=['GET'])
    def visionadmin_get_pages():
        locale = request.args.get('locale')
        include_deleted = request.args.get('trash') == '1'
        status_filter = request.args.get('status')
        query = (request.args.get('q') or '').strip()

        pages = Page.all(include_deleted=include_deleted)

        if include_deleted:
            pages = [p for p in pages if p.deleted_at is not None]
        else:
            pages = [p for p in pages if p.deleted_at is None]

        if status_filter == 'active':
            pages = [p for p in pages if p.is_active]
        elif status_filter == 'inactive':
            pages = [p for p in pages if not p.is_active]

        if query:
            q_lower = query.lower()
            pages = [
                p for p in pages
                if q_lower in p.get_title('en').lower()
                or q_lower in p.get_title('ar').lower()
                or q_lower in (p.slug or '').lower()
            ]

        all_active = [p for p in Page.all(include_deleted=False) if p.deleted_at is None]
        metrics = {
            'total': len(all_active),
            'active': len([p for p in all_active if p.is_active]),
            'inactive': len([p for p in all_active if not p.is_active]),
            'trash': len([p for p in Page.all(include_deleted=True) if p.deleted_at is not None])
        }

        return jsonify({
            'success': True,
            'pages': [p.to_dict(locale=locale) for p in pages],
            'metrics': metrics,
            'count': len(pages)
        })

    @app.route('/visionadmin/api/pages/<int:page_id>', methods=['GET'])
    @app.route('/visionadmin/api/v1/pages/<int:page_id>', methods=['GET'])
    def visionadmin_get_page(page_id):
        page = Page.find_by_id(page_id)
        if not page:
            return jsonify({'error': 'Page not found.'}), 404
        return jsonify({'success': True, 'page': page.to_dict()})

    @app.route('/visionadmin/api/pages', methods=['POST'])
    @app.route('/visionadmin/api/v1/pages', methods=['POST'])
    def visionadmin_create_page():
        data = request.get_json(silent=True) or {}

        # Validation
        title = data.get('title') or {}
        en_title = (title.get('en') if isinstance(title, dict) else str(title)).strip()
        if not en_title:
            return jsonify({'error': 'English Page Title is required.'}), 400

        slug = (data.get('slug') or '').strip()
        if not slug:
            slug = Page.slugify(en_title)
        else:
            slug = Page.slugify(slug)

        if not Page.is_slug_available(slug):
            return jsonify({'error': f'The slug "{slug}" is already in use. Please choose a different slug.'}), 409

        try:
            page = Page.create(
                title=title if isinstance(title, dict) else {"en": en_title, "ar": ""},
                slug=slug,
                content=data.get('content') or {"en": "", "ar": ""},
                banner_image=data.get('banner_image'),
                seo_title=data.get('seo_title'),
                meta_description=data.get('meta_description'),
                is_active=bool(data.get('is_active', True)),
                created_by=session.get('user_id'),
                updated_by=session.get('user_id')
            )
            return jsonify({
                'success': True,
                'page': page.to_dict(),
                'message': f'Page "{page.get_title()}" created successfully.'
            }), 201
        except Exception as e:
            if 'Duplicate entry' in str(e) or 'IntegrityError' in type(e).__name__:
                return jsonify({'error': f'The slug "{slug}" is already in use. Please enter a different slug.'}), 409
            return jsonify({'error': f'Failed to create page: {str(e)}'}), 500

    @app.route('/visionadmin/api/pages/<int:page_id>', methods=['PUT'])
    def visionadmin_update_page(page_id):
        page = Page.find_by_id(page_id)
        if not page:
            return jsonify({'error': 'Page not found.'}), 404

        data = request.get_json(silent=True) or {}

        if 'slug' in data and data['slug']:
            new_slug = Page.slugify(data['slug'])
            if not Page.is_slug_available(new_slug, exclude_id=page_id):
                return jsonify({'error': f'The slug "{new_slug}" is already in use.'}), 409
            data['slug'] = new_slug

        data['updated_by'] = session.get('user_id')

        try:
            page.update(**data)
            refreshed = Page.find_by_id(page_id)
            return jsonify({
                'success': True,
                'page': refreshed.to_dict(),
                'message': f'Page "{refreshed.get_title()}" updated successfully.'
            })
        except Exception as e:
            if 'Duplicate entry' in str(e) or 'IntegrityError' in type(e).__name__:
                return jsonify({'error': 'The specified slug is already in use.'}), 409
            return jsonify({'error': f'Failed to update page: {str(e)}'}), 500

    @app.route('/visionadmin/api/pages/<int:page_id>', methods=['DELETE'])
    def visionadmin_delete_page(page_id):
        page = Page.find_by_id(page_id)
        if not page:
            return jsonify({'error': 'Page not found.'}), 404

        is_hard = request.args.get('hard') == '1' or request.args.get('permanent') == '1' or page.deleted_at is not None

        if is_hard:
            Page.hard_delete(page_id)
            return jsonify({
                'success': True,
                'message': f'Page "{page.get_title()}" permanently deleted from database.'
            })
        else:
            Page.soft_delete(page_id)
            return jsonify({
                'success': True,
                'message': f'Page "{page.get_title()}" moved to trash.'
            })

    @app.route('/visionadmin/api/pages/<int:page_id>/restore', methods=['POST'])
    def visionadmin_restore_page(page_id):
        Page.restore(page_id)
        return jsonify({
            'success': True,
            'message': 'Page restored successfully.'
        })

    # =========================================================================
    # 3. BLOGS JSON API ENDPOINTS (/visionadmin/api/blogs & /visionadmin/api/v1/blogs)
    # =========================================================================

    @app.route('/visionadmin/api/categories', methods=['GET'])
    @app.route('/visionadmin/api/v1/categories', methods=['GET'])
    @app.route('/visionadmin/api/blog-categories', methods=['GET'])
    @app.route('/visionadmin/api/v1/blog-categories', methods=['GET'])
    def visionadmin_get_categories():
        """Returns all distinct category names from existing blogs table."""
        categories = Blog.distinct_categories()
        return jsonify({'success': True, 'categories': categories})

    @app.route('/visionadmin/api/blogs', methods=['GET'])
    @app.route('/visionadmin/api/v1/blogs', methods=['GET'])
    def visionadmin_get_blogs():
        locale = request.args.get('locale')
        include_deleted = request.args.get('trash') == '1'
        status_filter = request.args.get('status')
        query = (request.args.get('q') or '').strip()

        blogs = Blog.all(include_deleted=include_deleted)

        if include_deleted:
            blogs = [b for b in blogs if b.deleted_at is not None]
        else:
            blogs = [b for b in blogs if b.deleted_at is None]

        if status_filter in ('published', 'draft', 'archived'):
            blogs = [b for b in blogs if b.status == status_filter]

        if query:
            q_lower = query.lower()
            blogs = [
                b for b in blogs
                if q_lower in b.get_title('en').lower()
                or q_lower in b.get_title('ar').lower()
                or q_lower in (b.slug or '').lower()
                or q_lower in b.get_short_desc('en').lower()
                or q_lower in (b.category_name or '').lower()
            ]

        all_active = [b for b in Blog.all(include_deleted=False) if b.deleted_at is None]
        metrics = {
            'total': len(all_active),
            'published': len([b for b in all_active if b.status == 'published']),
            'draft': len([b for b in all_active if b.status == 'draft']),
            'archived': len([b for b in all_active if b.status == 'archived']),
            'trash': len([b for b in Blog.all(include_deleted=True) if b.deleted_at is not None])
        }

        return jsonify({
            'success': True,
            'blogs': [b.to_dict(locale=locale) for b in blogs],
            'metrics': metrics,
            'count': len(blogs)
        })

    @app.route('/visionadmin/api/blogs/<int:blog_id>', methods=['GET'])
    @app.route('/visionadmin/api/v1/blogs/<int:blog_id>', methods=['GET'])
    def visionadmin_get_blog(blog_id):
        blog = Blog.find_by_id(blog_id)
        if not blog:
            return jsonify({'error': 'Blog article not found.'}), 404
        return jsonify({'success': True, 'blog': blog.to_dict()})

    @app.route('/visionadmin/api/blogs', methods=['POST'])
    @app.route('/visionadmin/api/v1/blogs', methods=['POST'])
    def visionadmin_create_blog():
        data = request.get_json(silent=True) or {}

        title = data.get('title') or {}
        en_title = (title.get('en') if isinstance(title, dict) else str(title)).strip()
        if not en_title:
            return jsonify({'error': 'English Blog Title is required.'}), 400

        slug = (data.get('slug') or '').strip()
        if not slug:
            slug = Blog.slugify(en_title)
        else:
            slug = Blog.slugify(slug)

        if not Blog.is_slug_available(slug):
            return jsonify({'error': f'The slug "{slug}" is already in use. Please choose a unique slug.'}), 409

        try:
            blog = Blog.create(
                title=title if isinstance(title, dict) else {"en": en_title, "ar": ""},
                slug=slug,
                content=data.get('content') or {"en": "", "ar": ""},
                short_description=data.get('short_description') or {"en": "", "ar": ""},
                image=data.get('image'),
                category_name=(data.get('category_name') or data.get('category') or '').strip() or None,
                blog_category_id=data.get('blog_category_id'),
                author_id=data.get('author_id') or session.get('user_id') or 1,
                status=data.get('status') or 'draft',
                published_at=data.get('published_at'),
                meta_title=data.get('meta_title'),
                meta_desc=data.get('meta_desc'),
                faqs=data.get('faqs') or [],
                created_by=session.get('user_id'),
                updated_by=session.get('user_id')
            )
            return jsonify({
                'success': True,
                'blog': blog.to_dict(),
                'message': f'Blog "{blog.get_title()}" created successfully.'
            }), 201
        except Exception as e:
            if 'Duplicate entry' in str(e) or 'IntegrityError' in type(e).__name__:
                return jsonify({'error': f'The slug "{slug}" is already in use.'}), 409
            return jsonify({'error': f'Failed to create blog: {str(e)}'}), 500

    @app.route('/visionadmin/api/blogs/<int:blog_id>', methods=['PUT'])
    def visionadmin_update_blog(blog_id):
        blog = Blog.find_by_id(blog_id)
        if not blog:
            return jsonify({'error': 'Blog article not found.'}), 404

        data = request.get_json(silent=True) or {}

        if 'slug' in data and data['slug']:
            new_slug = Blog.slugify(data['slug'])
            if not Blog.is_slug_available(new_slug, exclude_id=blog_id):
                return jsonify({'error': f'The slug "{new_slug}" is already in use.'}), 409
            data['slug'] = new_slug

        data['updated_by'] = session.get('user_id')

        try:
            blog.update(**data)
            refreshed = Blog.find_by_id(blog_id)
            return jsonify({
                'success': True,
                'blog': refreshed.to_dict(),
                'message': f'Blog "{refreshed.get_title()}" updated successfully.'
            })
        except Exception as e:
            if 'Duplicate entry' in str(e) or 'IntegrityError' in type(e).__name__:
                return jsonify({'error': 'The specified slug is already in use.'}), 409
            return jsonify({'error': f'Failed to update blog: {str(e)}'}), 500

    @app.route('/visionadmin/api/blogs/<int:blog_id>', methods=['DELETE'])
    def visionadmin_delete_blog(blog_id):
        blog = Blog.find_by_id(blog_id)
        if not blog:
            return jsonify({'error': 'Blog article not found.'}), 404

        is_hard = request.args.get('hard') == '1' or request.args.get('permanent') == '1' or blog.deleted_at is not None

        if is_hard:
            Blog.hard_delete(blog_id)
            return jsonify({
                'success': True,
                'message': f'Article "{blog.get_title()}" permanently deleted from database.'
            })
        else:
            Blog.soft_delete(blog_id)
            return jsonify({
                'success': True,
                'message': f'Article "{blog.get_title()}" moved to trash.'
            })

    @app.route('/visionadmin/api/blogs/<int:blog_id>/restore', methods=['POST'])
    def visionadmin_restore_blog(blog_id):
        Blog.restore(blog_id)
        return jsonify({
            'success': True,
            'message': 'Blog article restored successfully.'
        })

    # =========================================================================
    # 4. PAGE SECTIONS CRUD & REORDER API (/visionadmin/api/sections & /visionadmin/api/v1/sections)
    # =========================================================================

    @app.route('/visionadmin/api/sections', methods=['GET'])
    @app.route('/visionadmin/api/v1/sections', methods=['GET'])
    def visionadmin_get_sections():
        """Returns all sections for a page (including inactive) ordered by sort_order."""
        page_slug = request.args.get('page') or request.args.get('page_slug') or 'about-us'
        sections = PageSection.all_for_page(page_slug=page_slug, include_inactive=True)
        return jsonify({
            'page': page_slug,
            'sections': sections,
            'count': len(sections)
        })

    @app.route('/visionadmin/api/sections/<int:section_id>', methods=['GET'])
    @app.route('/visionadmin/api/v1/sections/<int:section_id>', methods=['GET'])
    def visionadmin_get_section_detail(section_id):
        """Returns a single section by id."""
        sec = PageSection.find_by_id(section_id)
        if not sec:
            return jsonify({'error': 'Section not found.'}), 404
        return jsonify({'section': sec})

    @app.route('/visionadmin/api/sections', methods=['POST'])
    @app.route('/visionadmin/api/v1/sections', methods=['POST'])
    def visionadmin_create_section():
        """Creates a new section."""
        data = request.get_json(silent=True) or request.form.to_dict()
        if not data:
            return jsonify({'error': 'No data provided.'}), 400

        section_type = data.get('section_type')
        if not section_type:
            return jsonify({'error': 'section_type is required.'}), 400

        try:
            new_sec = PageSection.create(data)
            return jsonify({
                'success': True,
                'section': new_sec,
                'message': 'Section added successfully.'
            }), 201
        except Exception as e:
            return jsonify({'error': f'Failed to create section: {str(e)}'}), 500

    @app.route('/visionadmin/api/sections/<int:section_id>', methods=['PUT'])
    def visionadmin_update_section(section_id):
        """Updates an existing section."""
        sec = PageSection.find_by_id(section_id)
        if not sec:
            return jsonify({'error': 'Section not found.'}), 404

        data = request.get_json(silent=True) or request.form.to_dict()
        if not data:
            return jsonify({'error': 'No data provided.'}), 400

        try:
            updated = PageSection.update(section_id, data)
            return jsonify({
                'success': True,
                'section': updated,
                'message': 'Section updated successfully.'
            })
        except Exception as e:
            return jsonify({'error': f'Failed to update section: {str(e)}'}), 500

    @app.route('/visionadmin/api/sections/<int:section_id>', methods=['DELETE'])
    def visionadmin_delete_section(section_id):
        """Soft deletes a section."""
        sec = PageSection.find_by_id(section_id)
        if not sec:
            return jsonify({'error': 'Section not found.'}), 404

        PageSection.soft_delete(section_id)
        return jsonify({
            'success': True,
            'message': 'Section deleted successfully.'
        })

    @app.route('/visionadmin/api/sections/<int:section_id>/toggle', methods=['POST'])
    def visionadmin_toggle_section(section_id):
        """Toggles active/disabled state of a section."""
        sec = PageSection.find_by_id(section_id)
        if not sec:
            return jsonify({'error': 'Section not found.'}), 404

        toggled = PageSection.toggle_active(section_id)
        status_str = 'enabled' if toggled.get('is_active') else 'disabled'
        return jsonify({
            'success': True,
            'section': toggled,
            'message': f'Section {status_str} successfully.'
        })

    @app.route('/visionadmin/api/sections/reorder', methods=['POST'])
    def visionadmin_reorder_sections():
        """Updates section order based on ordered list of IDs."""
        data = request.get_json(silent=True) or {}
        ordered_ids = data.get('ordered_ids') or data.get('ids') or []
        if not ordered_ids or not isinstance(ordered_ids, list):
            return jsonify({'error': 'ordered_ids list is required.'}), 400

        try:
            PageSection.reorder(ordered_ids)
            return jsonify({
                'success': True,
                'message': 'Section order saved successfully.'
            })
        except Exception as e:
            return jsonify({'error': f'Failed to reorder sections: {str(e)}'}), 500

    # =========================================================================
    # 5. REVIEWER SETTINGS CONFIGURATION API
    # =========================================================================

    @app.route('/visionadmin/api/settings/reviewer', methods=['GET'])
    @app.route('/visionadmin/api/v1/settings/reviewer', methods=['GET'])
    @app.route('/visionadmin/api/reviewer-settings', methods=['GET'])
    @app.route('/visionadmin/api/v1/reviewer-settings', methods=['GET'])
    def visionadmin_get_reviewer_settings():
        from models.setting import Setting
        settings = Setting.get_reviewer_settings()
        return jsonify({
            'success': True,
            'settings': settings
        })

    @app.route('/visionadmin/api/settings/reviewer', methods=['POST', 'PUT'])
    @app.route('/visionadmin/api/v1/settings/reviewer', methods=['POST', 'PUT'])
    @app.route('/visionadmin/api/reviewer-settings', methods=['POST', 'PUT'])
    @app.route('/visionadmin/api/v1/reviewer-settings', methods=['POST', 'PUT'])
    def visionadmin_save_reviewer_settings():
        from models.setting import Setting
        data = request.get_json(silent=True) or request.form.to_dict() or {}
        
        # Normalize enabled flag
        enabled_val = data.get('enabled', True)
        if isinstance(enabled_val, str):
            enabled = enabled_val.strip().lower() in ('true', '1', 'yes')
        else:
            enabled = bool(enabled_val)

        payload = {
            'enabled': enabled,
            'name': data.get('name') if isinstance(data.get('name'), dict) else {
                'en': (data.get('name_en') or data.get('name') or 'Sharvil Kumar').strip(),
                'ar': (data.get('name_ar') or 'شارفيل كومار').strip()
            },
            'initials': (data.get('initials') or 'SK').strip(),
            'role': data.get('role') if isinstance(data.get('role'), dict) else {
                'en': (data.get('role_en') or data.get('role') or 'Tyre Selection Specialist, TyresCart').strip(),
                'ar': (data.get('role_ar') or 'أخصائي اختيار الإطارات، تايرز كارت').strip()
            },
            'bio': data.get('bio') if isinstance(data.get('bio'), dict) else {
                'en': (data.get('bio_en') or data.get('bio') or data.get('description_en') or '').strip(),
                'ar': (data.get('bio_ar') or data.get('description_ar') or '').strip()
            }
        }

        Setting.set('reviewer_settings', payload, group='reviewer')
        return jsonify({
            'success': True,
            'settings': Setting.get_reviewer_settings(),
            'message': 'Reviewer settings saved successfully.'
        })

    # =========================================================================
    # 6. UNIFIED GLOBAL SEARCH API (Deep search across all CMS content & sections)
    # =========================================================================

    @app.route('/visionadmin/api/global-search', methods=['GET'])
    @app.route('/visionadmin/api/v1/global-search', methods=['GET'])
    def visionadmin_global_search():
        """
        Deep content search across:
        1. Pages (title, slug, content HTML/prose, meta_description, seo_title)
        2. Page Sections (section_title, section_subtitle, content, page_slug, section_type)
        3. Blogs & Articles (title, content, slug, category, short_description)
        """
        query = (request.args.get('q') or '').strip()
        if not query:
            return jsonify({
                'success': True,
                'query': '',
                'total': 0,
                'results': {'pages': [], 'sections': [], 'blogs': []}
            })

        q_lower = query.lower()

        def clean_html(text):
            if not text:
                return ''
            clean = re.sub(r'<[^>]+>', ' ', str(text))
            return ' '.join(clean.split())

        def make_snippet(text, q, max_len=110):
            cleaned = clean_html(text)
            idx = cleaned.lower().find(q)
            if idx == -1:
                return cleaned[:max_len] + ('...' if len(cleaned) > max_len else '')
            start = max(0, idx - 25)
            end = min(len(cleaned), idx + len(q) + 55)
            snippet = cleaned[start:end]
            if start > 0:
                snippet = '...' + snippet
            if end < len(cleaned):
                snippet = snippet + '...'
            return snippet

        # 1. Search Pages
        all_pages = [p for p in Page.all(include_deleted=False) if p.deleted_at is None]
        matched_pages = []
        for p in all_pages:
            title_en = p.get_title('en')
            title_ar = p.get_title('ar')
            slug = p.slug or ''
            content_en = clean_html(p.get_content('en'))
            content_ar = clean_html(p.get_content('ar'))
            meta_desc = clean_html(p.get_meta_desc('en'))

            match_found = False
            snippet = ''
            if q_lower in title_en.lower() or q_lower in title_ar.lower():
                match_found = True
                snippet = title_en
            elif q_lower in slug.lower():
                match_found = True
                snippet = f"/{slug.lstrip('/')}"
            elif q_lower in content_en.lower():
                match_found = True
                snippet = make_snippet(content_en, q_lower)
            elif q_lower in content_ar.lower():
                match_found = True
                snippet = make_snippet(content_ar, q_lower)
            elif q_lower in meta_desc.lower():
                match_found = True
                snippet = make_snippet(meta_desc, q_lower)

            if match_found:
                matched_pages.append({
                    'id': p.id,
                    'type': 'page',
                    'title': title_en or slug,
                    'slug': f"/{slug.lstrip('/')}",
                    'snippet': snippet,
                    'is_active': bool(p.is_active),
                    'url': f"/visionadmin/pages#page-{p.id}"
                })

        # 2. Search Page Sections
        conn = get_connection()
        matched_sections = []
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT id, page_slug, section_type, section_title, section_subtitle, content, is_active
                    FROM page_sections
                    WHERE deleted_at IS NULL
                    ORDER BY sort_order ASC, id ASC
                """)
                sec_rows = cursor.fetchall() or []
                for row in sec_rows:
                    sec_id = row.get('id')
                    page_slug = row.get('page_slug') or 'about-us'
                    sec_type = row.get('section_type') or 'section'
                    sec_title_raw = PageSection._parse_json(row.get('section_title'))
                    sec_title_en = sec_title_raw.get('en') if isinstance(sec_title_raw, dict) else str(sec_title_raw or '')
                    sec_sub_raw = PageSection._parse_json(row.get('section_subtitle'))
                    sec_sub_en = sec_sub_raw.get('en') if isinstance(sec_sub_raw, dict) else str(sec_sub_raw or '')
                    sec_content_raw = PageSection._parse_json(row.get('content'))
                    sec_content_en = clean_html(sec_content_raw.get('en') if isinstance(sec_content_raw, dict) else str(sec_content_raw or ''))

                    match_found = False
                    snippet = ''
                    if q_lower in sec_title_en.lower():
                        match_found = True
                        snippet = sec_title_en
                    elif q_lower in sec_sub_en.lower():
                        match_found = True
                        snippet = sec_sub_en
                    elif q_lower in sec_content_en.lower():
                        match_found = True
                        snippet = make_snippet(sec_content_en, q_lower)
                    elif q_lower in page_slug.lower() or q_lower in sec_type.lower():
                        match_found = True
                        snippet = f"Page: {page_slug} ({sec_type})"

                    if match_found:
                        display_title = sec_title_en or f"{page_slug.replace('-', ' ').title()} {sec_type.title()} Section"
                        matched_sections.append({
                            'id': sec_id,
                            'type': 'section',
                            'title': display_title,
                            'slug': f"/en/{page_slug.lstrip('/')} ({sec_type})",
                            'page_slug': page_slug,
                            'section_type': sec_type,
                            'snippet': snippet,
                            'is_active': bool(row.get('is_active', 1)),
                            'url': f"/visionadmin/sections?page={page_slug}#section-{sec_id}"
                        })
        finally:
            conn.close()

        # 3. Search Blogs & Articles
        all_blogs = [b for b in Blog.all(include_deleted=False) if b.deleted_at is None]
        matched_blogs = []
        for b in all_blogs:
            title_en = b.get_title('en')
            slug = b.slug or ''
            content_en = clean_html(b.get_content('en'))
            cat_name = b.category_name or ''

            match_found = False
            snippet = ''
            if q_lower in title_en.lower():
                match_found = True
                snippet = title_en
            elif q_lower in slug.lower():
                match_found = True
                snippet = f"/blog/{slug}"
            elif q_lower in cat_name.lower():
                match_found = True
                snippet = f"Category: {cat_name}"
            elif q_lower in content_en.lower():
                match_found = True
                snippet = make_snippet(content_en, q_lower)

            if match_found:
                matched_blogs.append({
                    'id': b.id,
                    'type': 'blog',
                    'title': title_en or slug,
                    'slug': f"/blog/{slug.lstrip('/')}",
                    'category': cat_name,
                    'snippet': snippet,
                    'is_active': b.status == 'published',
                    'url': f"/visionadmin/blogs#blog-{b.id}"
                })

        total = len(matched_pages) + len(matched_sections) + len(matched_blogs)
        return jsonify({
            'success': True,
            'query': query,
            'total': total,
            'results': {
                'pages': matched_pages[:5],
                'sections': matched_sections[:5],
                'blogs': matched_blogs[:5]
            }
        })

    # =========================================================================
    # 7. ENQUIRIES / LEADS MANAGEMENT API (hdweb_enquiry table)
    # =========================================================================

    @app.route('/visionadmin/api/enquiries', methods=['GET'])
    @app.route('/visionadmin/api/v1/enquiries', methods=['GET'])
    def visionadmin_get_enquiries():
        """
        Fetches all enquiries from hdweb_enquiry with metrics, filtering, and search.
        """
        q = (request.args.get('q') or '').strip().lower()
        status_filter = request.args.get('status')
        form_type_filter = request.args.get('form_type')

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "SELECT * FROM hdweb_enquiry ORDER BY enquiry_id DESC"
                cursor.execute(sql)
                rows = cursor.fetchall() or []

                total_count = len(rows)
                new_count = 0
                banner_count = 0
                wa_count = 0

                filtered = []
                for r in rows:
                    st = r.get('status', 0)
                    ft = (r.get('form_type') or '').lower()

                    if st == 0:
                        new_count += 1
                    if 'banner' in ft:
                        banner_count += 1
                    else:
                        wa_count += 1

                    if status_filter is not None and status_filter != '' and status_filter != 'all':
                        try:
                            if int(st) != int(status_filter):
                                continue
                        except (ValueError, TypeError):
                            pass

                    if form_type_filter and form_type_filter != 'all':
                        if form_type_filter == 'banner' and 'banner' not in ft:
                            continue
                        elif form_type_filter == 'whatsapp' and 'banner' in ft:
                            continue

                    if q:
                        haystack = " ".join(str(v or '') for v in [
                            r.get('name'), r.get('email'), r.get('number'),
                            r.get('vehicle'), r.get('make'), r.get('model'),
                            r.get('tyre_size'), r.get('city'), r.get('message'),
                            r.get('enquiry_for')
                        ]).lower()
                        if q not in haystack:
                            continue

                    dt = r.get('created_at')
                    created_at_fmt = to_ist_12h(dt, with_seconds=False) if dt else None
                    created_at_raw = dt.isoformat() + 'Z' if dt and hasattr(dt, 'isoformat') else None

                    status_map = {0: 'New', 1: 'In Progress', 2: 'Resolved', 3: 'Closed'}
                    st_val = r.get('status', 0)

                    filtered.append({
                        'enquiry_id': r.get('enquiry_id'),
                        'name': r.get('name'),
                        'email': r.get('email'),
                        'number': r.get('number'),
                        'enquiry_for': r.get('enquiry_for'),
                        'message': r.get('message'),
                        'status': st_val,
                        'status_label': status_map.get(st_val, 'New'),
                        'form_type': r.get('form_type'),
                        'model': r.get('model'),
                        'make': r.get('make'),
                        'year': r.get('year'),
                        'spec': r.get('spec'),
                        'current_insurance': r.get('current_insurance'),
                        'vehicle': r.get('vehicle'),
                        'tyre_size': r.get('tyre_size'),
                        'city': r.get('city'),
                        'created_at': created_at_fmt,
                        'created_at_raw': created_at_raw
                    })

                return jsonify({
                    'success': True,
                    'enquiries': filtered,
                    'metrics': {
                        'total': total_count,
                        'new': new_count,
                        'banner': banner_count,
                        'whatsapp_direct': wa_count
                    },
                    'count': len(filtered)
                })
        finally:
            conn.close()

    @app.route('/visionadmin/api/enquiries/<int:enquiry_id>', methods=['GET'])
    @app.route('/visionadmin/api/v1/enquiries/<int:enquiry_id>', methods=['GET'])
    def visionadmin_get_single_enquiry(enquiry_id):
        """Fetches detail of a single enquiry by enquiry_id."""
        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT * FROM hdweb_enquiry WHERE enquiry_id = %s", (enquiry_id,))
                row = cursor.fetchone()
                if not row:
                    return jsonify({'success': False, 'error': 'Enquiry not found'}), 404

                dt = row.get('created_at')
                created_at_fmt = to_ist_12h(dt, with_seconds=False) if dt else None
                created_at_raw = dt.isoformat() + 'Z' if dt and hasattr(dt, 'isoformat') else None
                status_map = {0: 'New', 1: 'In Progress', 2: 'Resolved', 3: 'Closed'}

                item = {**row}
                item['created_at'] = created_at_fmt
                item['created_at_raw'] = created_at_raw
                item['status_label'] = status_map.get(row.get('status', 0), 'New')

                return jsonify({'success': True, 'enquiry': item})
        finally:
            conn.close()

    @app.route('/visionadmin/api/enquiries/<int:enquiry_id>/status', methods=['PUT', 'POST'])
    @app.route('/visionadmin/api/v1/enquiries/<int:enquiry_id>/status', methods=['PUT', 'POST'])
    def visionadmin_update_enquiry_status(enquiry_id):
        """Updates the status of an enquiry."""
        data = request.get_json(silent=True) or request.form.to_dict() or {}
        new_status = data.get('status')
        if new_status is None:
            return jsonify({'success': False, 'error': 'Status is required'}), 400

        try:
            status_int = int(new_status)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid status value'}), 400

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("UPDATE hdweb_enquiry SET status = %s WHERE enquiry_id = %s", (status_int, enquiry_id))
                conn.commit()
                return jsonify({'success': True, 'message': 'Status updated successfully', 'status': status_int})
        finally:
            conn.close()

    # =========================================================================
    # 8. VISIONADMIN USER MANAGEMENT API (admin_users table)
    # =========================================================================

    @app.route('/visionadmin/api/users', methods=['GET'])
    @app.route('/visonadmin/api/users', methods=['GET'])
    def visionadmin_api_list_users():
        from visionadmin.admin_auth import list_admin_users, get_admin_user_metrics
        is_trash = request.args.get('trash') in ('1', 'true', 'yes')
        users = list_admin_users(is_trash=is_trash)
        metrics = get_admin_user_metrics()

        return jsonify({
            'success': True,
            'users': users,
            'is_trash': is_trash,
            'metrics': {
                'total': metrics.get('total', 0),
                'super_admins': metrics.get('super', 0),
                'managers': metrics.get('managers', 0),
                'active': metrics.get('active', 0),
                'trash': metrics.get('trash', 0)
            }
        })

    @app.route('/visionadmin/api/users', methods=['POST'])
    @app.route('/visonadmin/api/users', methods=['POST'])
    def visionadmin_api_create_user():
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can create admin users.'}), 403

        data = request.get_json(silent=True) or request.form
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip().lower()
        password = data.get('password') or ''
        role = (data.get('role') or 'manager').strip().lower()
        is_active = 1 if data.get('is_active') in (1, True, '1', 'true', 'on') else 0

        if not name:
            return jsonify({'error': 'Full name is required.'}), 400
        if not email or not EMAIL_RE.match(email):
            return jsonify({'error': 'A valid email address is required.'}), 400
        if not password:
            password = secrets.token_urlsafe(16)
        elif len(password) < 8:
            return jsonify({'error': 'Password must be at least 8 characters long.'}), 400
        if role not in ('super_admin', 'manager', 'support'):
            return jsonify({'error': 'Invalid role. Choose Super Admin, Manager, or Support.'}), 400

        from visionadmin.admin_auth import create_admin_user, get_admin_user_by_email, get_admin_user_by_id, serialize_admin_user

        existing = get_admin_user_by_email(email)
        if existing:
            return jsonify({'error': f"An administrator account with email '{email}' already exists."}), 409

        try:
            new_id = create_admin_user(name, email, password, role, is_active)
            created = get_admin_user_by_id(new_id)

            # Send welcome email with login credentials and reset password link
            try:
                from mailer import send_email
                from visionadmin.admin_auth import create_admin_password_reset_token

                email_img_dir = os.path.join(app.static_folder or 'static', 'assets', 'images', 'email')
                logo_file = os.path.join(app.static_folder or 'static', 'assets', 'images', 'tyresvision-email-logo.png')
                
                all_cids = {
                    'tyresvision_logo': logo_file,
                    'hero_shield_badge': os.path.join(email_img_dir, 'hero_shield_badge.png'),
                    'badge_user': os.path.join(email_img_dir, 'badge_user.png'),
                    'badge_lock': os.path.join(email_img_dir, 'badge_lock.png'),
                    'badge_tip': os.path.join(email_img_dir, 'badge_tip.png'),
                    'row_user': os.path.join(email_img_dir, 'row_user.png'),
                    'row_email': os.path.join(email_img_dir, 'row_email.png'),
                    'social_fb': os.path.join(email_img_dir, 'social_fb.png'),
                    'social_in': os.path.join(email_img_dir, 'social_in.png'),
                    'social_ig': os.path.join(email_img_dir, 'social_ig.png'),
                    'social_yt': os.path.join(email_img_dir, 'social_yt.png'),
                }
                inline_imgs = {k: v for k, v in all_cids.items() if os.path.isfile(v)}
                
                token = create_admin_password_reset_token(email)
                reset_link = f"{request.host_url.rstrip('/')}/visionadmin/reset-password?token={token}"
                login_link = f"{request.host_url.rstrip('/')}/visionadmin/login?email={email}"
                logo_url = f"{request.host_url.rstrip('/')}/static/assets/images/tyresvision-email-logo.png"
                
                html_body = render_template(
                    'emails/welcome_user.html',
                    user_name=name,
                    user_email=email,
                    user_role=role,
                    reset_link=reset_link,
                    login_link=login_link,
                    logo_cid='tyresvision_logo' if 'tyresvision_logo' in inline_imgs else None,
                    hero_shield_cid='hero_shield_badge' if 'hero_shield_badge' in inline_imgs else None,
                    badge_user_cid='badge_user' if 'badge_user' in inline_imgs else None,
                    badge_lock_cid='badge_lock' if 'badge_lock' in inline_imgs else None,
                    badge_tip_cid='badge_tip' if 'badge_tip' in inline_imgs else None,
                    row_user_cid='row_user' if 'row_user' in inline_imgs else None,
                    row_email_cid='row_email' if 'row_email' in inline_imgs else None,
                    social_fb_cid='social_fb' if 'social_fb' in inline_imgs else None,
                    social_in_cid='social_in' if 'social_in' in inline_imgs else None,
                    social_ig_cid='social_ig' if 'social_ig' in inline_imgs else None,
                    social_yt_cid='social_yt' if 'social_yt' in inline_imgs else None,
                    logo_url=logo_url,
                )
                send_email(
                    email,
                    'Welcome to TyresVision! Your Account Details',
                    html_body,
                    inline_images=inline_imgs,
                )
            except Exception as mail_err:
                app.logger.warning(f"Failed to send welcome email to {email}: {mail_err}")

            return jsonify({
                'success': True,
                'message': f"Administrator '{name}' created successfully. Welcome email sent.",
                'user': serialize_admin_user(created) if created else {}
            }), 201
        except Exception as err:
            return jsonify({'error': f"Failed to create admin user: {err}"}), 500

    @app.route('/visionadmin/api/users/<int:user_id>', methods=['GET'])
    @app.route('/visonadmin/api/users/<int:user_id>', methods=['GET'])
    def visionadmin_api_get_user(user_id):
        from visionadmin.admin_auth import get_admin_user_by_id, serialize_admin_user
        user = get_admin_user_by_id(user_id)
        if not user:
            return jsonify({'error': 'Administrator user not found.'}), 404
        return jsonify({'success': True, 'user': serialize_admin_user(user)})

    @app.route('/visionadmin/api/users/<int:user_id>', methods=['PUT'])
    @app.route('/visonadmin/api/users/<int:user_id>', methods=['PUT'])
    def visionadmin_api_update_user(user_id):
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can update admin users.'}), 403

        from visionadmin.admin_auth import get_admin_user_by_id, update_admin_user, get_admin_user_by_email, serialize_admin_user, count_super_admins

        existing = get_admin_user_by_id(user_id)
        if not existing:
            return jsonify({'error': 'Administrator user not found.'}), 404

        data = request.get_json(silent=True) or request.form
        name = (data.get('name') or existing['name']).strip()
        email = (data.get('email') or existing['email']).strip().lower()
        role = (data.get('role') or existing['role']).strip().lower()
        is_active_val = data.get('is_active')
        is_active = 1 if is_active_val in (1, True, '1', 'true', 'on') else (0 if is_active_val in (0, False, '0', 'false') else existing.get('is_active', 1))
        password = data.get('password')

        if not name:
            return jsonify({'error': 'Name cannot be empty.'}), 400
        if not email or not EMAIL_RE.match(email):
            return jsonify({'error': 'A valid email address is required.'}), 400
        if role not in ('super_admin', 'manager', 'support'):
            return jsonify({'error': 'Invalid role.'}), 400
        if password and len(password) < 8:
            return jsonify({'error': 'New password must be at least 8 characters.'}), 400

        # Check if demoting the only super_admin
        if existing.get('role') == 'super_admin' and role != 'super_admin' and count_super_admins() <= 1:
            return jsonify({'error': 'Action blocked. Cannot demote the only remaining Super Administrator.'}), 400

        # Check email collision
        if email != existing['email'].strip().lower():
            dup = get_admin_user_by_email(email)
            if dup and dup['id'] != user_id:
                return jsonify({'error': f"Email '{email}' is already in use."}), 409

        try:
            update_admin_user(user_id, name, email, role, is_active, password if password else None)
            updated = get_admin_user_by_id(user_id)
            return jsonify({
                'success': True,
                'message': f"Administrator '{name}' updated successfully.",
                'user': serialize_admin_user(updated)
            })
        except Exception as err:
            return jsonify({'error': f"Failed to update user: {err}"}), 500

    @app.route('/visionadmin/api/users/<int:user_id>', methods=['DELETE'])
    @app.route('/visonadmin/api/users/<int:user_id>', methods=['DELETE'])
    def visionadmin_api_delete_user(user_id):
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can delete admin users.'}), 403

        current_admin_id = session.get('admin_user_id') or session.get('user_id')
        if user_id == current_admin_id:
            return jsonify({'error': 'Action not allowed. You cannot delete your own active account.'}), 400

        from visionadmin.admin_auth import get_admin_user_by_id, delete_admin_user, count_super_admins

        target = get_admin_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'Administrator user not found.'}), 404

        if target.get('role') == 'super_admin' and count_super_admins() <= 1:
            return jsonify({'error': 'Action blocked. Cannot delete the only remaining Super Administrator.'}), 400

        try:
            delete_admin_user(user_id)
            return jsonify({
                'success': True,
                'message': f"Administrator account '{target['name']}' moved to trash."
            })
        except Exception as err:
            return jsonify({'error': f"Failed to move user to trash: {err}"}), 500

    @app.route('/visionadmin/api/users/<int:user_id>/restore', methods=['POST'])
    @app.route('/visonadmin/api/users/<int:user_id>/restore', methods=['POST'])
    def visionadmin_api_restore_user(user_id):
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can restore admin users.'}), 403

        from visionadmin.admin_auth import get_admin_user_by_id, restore_admin_user

        target = get_admin_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'Administrator user not found.'}), 404

        try:
            restore_admin_user(user_id)
            return jsonify({
                'success': True,
                'message': f"Administrator account '{target['name']}' restored from trash."
            })
        except Exception as err:
            return jsonify({'error': f"Failed to restore user: {err}"}), 500

    @app.route('/visionadmin/api/users/<int:user_id>/purge', methods=['DELETE', 'POST'])
    @app.route('/visonadmin/api/users/<int:user_id>/purge', methods=['DELETE', 'POST'])
    def visionadmin_api_purge_user(user_id):
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can permanently delete admin users.'}), 403

        current_admin_id = session.get('admin_user_id') or session.get('user_id')
        if user_id == current_admin_id:
            return jsonify({'error': 'Action not allowed. You cannot delete your own account.'}), 400

        from visionadmin.admin_auth import get_admin_user_by_id, permanent_delete_admin_user

        target = get_admin_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'Administrator user not found.'}), 404

        try:
            permanent_delete_admin_user(user_id)
            return jsonify({
                'success': True,
                'message': f"Administrator account '{target['name']}' permanently deleted."
            })
        except Exception as err:
            return jsonify({'error': f"Failed to permanently delete user: {err}"}), 500

    @app.route('/visionadmin/api/users/<int:user_id>/toggle-status', methods=['POST'])
    @app.route('/visonadmin/api/users/<int:user_id>/toggle-status', methods=['POST'])
    def visionadmin_api_toggle_user_status(user_id):
        role_norm = str(session.get('role') or '').strip().lower().replace('-', '_').replace(' ', '_')
        if role_norm not in ('super_admin', 'superadmin') and session.get('role') != 'SuperAdmin':
            return jsonify({'error': 'Forbidden. Only Super Administrators can modify account status.'}), 403

        current_admin_id = session.get('admin_user_id') or session.get('user_id')
        if user_id == current_admin_id:
            return jsonify({'error': 'Action not allowed. You cannot disable your own active account.'}), 400

        from visionadmin.admin_auth import get_admin_user_by_id, toggle_admin_user_status, count_super_admins

        target = get_admin_user_by_id(user_id)
        if not target:
            return jsonify({'error': 'Administrator user not found.'}), 404

        if target.get('role') == 'super_admin' and target.get('is_active', 1) == 1 and count_super_admins() <= 1:
            return jsonify({'error': 'Action blocked. Cannot disable the only remaining active Super Administrator.'}), 400

        new_status = toggle_admin_user_status(user_id)
        return jsonify({
            'success': True,
            'is_active': new_status,
            'message': f"Administrator account '{target['name']}' {'activated' if new_status else 'disabled'}."
        })


def register_client_api_routes(app):
    """Registers all public, un-prefixed /api/* endpoints for the client storefront."""

    # =========================================================================
    # 1. BLOG JSON API (GET BLOG FROM DATABASE)
    # =========================================================================

    @app.route('/api/blogs', methods=['GET'])
    @app.route('/api/v1/blogs', methods=['GET'])
    @app.route('/api/blog', methods=['GET'])
    @app.route('/api/v1/blog', methods=['GET'])
    def api_get_blogs():
        """
        Public JSON API: Fetch published blogs with pagination, locale,
        search, and category filtering.
        """
        locale = request.args.get('locale') or _get_locale()
        query = (request.args.get('q') or '').strip().lower()
        cat_filter = (request.args.get('category') or '').strip().lower()

        try:
            page_num = max(1, int(request.args.get('page', 1)))
        except (ValueError, TypeError):
            page_num = 1

        try:
            per_page = int(request.args.get('per_page', request.args.get('limit', 12)))
            if per_page not in (4, 6, 8, 12, 16, 24):
                per_page = 12
        except (ValueError, TypeError):
            per_page = 12

        db_blogs = Blog.published()
        formatted_blogs = []

        if db_blogs:
            for b in db_blogs:
                title = b.get_title(locale)
                short_desc = b.get_short_desc(locale)
                content = b.get_content(locale)

                cat_name = b.category_name or ('Blog' if locale != 'ar' else 'مدونة')

                prefix = f'/{locale}' if locale in ('en', 'ar') else ''
                blog_url = f'{prefix}/blog/{b.slug}'

                formatted_blogs.append({
                    'id': b.id,
                    'slug': b.slug,
                    'title': title,
                    'short_description': short_desc or '',
                    'excerpt': short_desc or '',
                    'content': content or '',
                    'image': b.image or '/static/assets/online-tyres-shop-dubai.png',
                    'cover_image_url': b.image or '/static/assets/online-tyres-shop-dubai.png',
                    'published_at': b.published_at.strftime('%d-%m-%Y') if b.published_at else (b.created_at.strftime('%d-%m-%Y') if b.created_at else '2026'),
                    'published_at_raw': b.published_at.isoformat() if b.published_at else (b.created_at.isoformat() if b.created_at else None),
                    'category': cat_name,
                    'thumb_class': 't-buying' if 'choose' in (b.slug or '') else 't-maint',
                    'read_time': '4 min read' if locale != 'ar' else 'قراءة 4 دقائق',
                    'url': blog_url
                })

        # Filter by search keyword
        if query:
            formatted_blogs = [
                b for b in formatted_blogs
                if query in b['title'].lower() or query in b['short_description'].lower()
            ]

        # Filter by category
        if cat_filter:
            formatted_blogs = [
                b for b in formatted_blogs
                if cat_filter == (b.get('category') or '').strip().lower()
                or cat_filter in (b.get('category') or '').strip().lower()
                or cat_filter == Blog.slugify(b.get('category') or '')
            ]

        total_count = len(formatted_blogs)
        num_pages = max(1, math.ceil(total_count / per_page))
        if page_num > num_pages:
            page_num = num_pages

        start_idx = (page_num - 1) * per_page
        end_idx = min(start_idx + per_page, total_count)
        page_blogs = formatted_blogs[start_idx:end_idx]

        pagination = {
            'page': page_num,
            'per_page': per_page,
            'total': total_count,
            'num_pages': num_pages,
            'start': start_idx + 1 if total_count > 0 else 0,
            'end': end_idx,
            'has_prev': page_num > 1,
            'has_next': page_num < num_pages,
            'prev_num': page_num - 1,
            'next_num': page_num + 1
        }

        return jsonify({
            'success': True,
            'locale': locale,
            'blogs': page_blogs,
            'count': len(page_blogs),
            'pagination': pagination
        })

    @app.route('/api/blogs/<slug>', methods=['GET'])
    @app.route('/api/v1/blogs/<slug>', methods=['GET'])
    @app.route('/api/blog/<slug>', methods=['GET'])
    @app.route('/api/v1/blog/<slug>', methods=['GET'])
    def api_get_blog_detail(slug):
        """
        Public JSON API: Fetch a single blog article by slug.
        """
        locale = request.args.get('locale') or _get_locale()
        blog = Blog.find_by_slug(slug)

        if not blog:
            return jsonify({'success': False, 'error': 'Blog not found'}), 404

        data = {
            'id': blog.id,
            'slug': blog.slug,
            'title': blog.get_title(locale),
            'short_description': blog.get_short_desc(locale),
            'content': blog.get_content(locale),
            'image': blog.image or '/static/assets/online-tyres-shop-dubai.png',
            'cover_image_url': blog.image or '/static/assets/online-tyres-shop-dubai.png',
            'published_at': blog.published_at.strftime('%d-%m-%Y') if blog.published_at else '24-08-2026',
            'meta_title': blog.get_meta_title(locale),
            'meta_desc': blog.get_meta_desc(locale),
            'author': {
                'name': 'Sharvil Kumar' if locale != 'ar' else 'شارفيل كومار',
                'role': 'Tyre Selection Specialist, TyresVision' if locale != 'ar' else 'أخصائي اختيار الإطارات، تايرز فيجن',
                'avatar_initials': 'SK'
            }
        }
        return jsonify({'success': True, 'blog': data})

    # =========================================================================
    # 2. PUBLIC PAGE SECTIONS API (Returns active sections ordered by sort_order)
    # =========================================================================

    @app.route('/api/pages/<slug>/sections', methods=['GET'])
    @app.route('/api/v1/pages/<slug>/sections', methods=['GET'])
    @app.route('/api/sections/<slug>', methods=['GET'])
    @app.route('/api/v1/sections/<slug>', methods=['GET'])
    @app.route('/api/sections', methods=['GET'])
    @app.route('/api/v1/sections', methods=['GET'])
    @app.route('/api/pages/about-us/sections', methods=['GET'])
    @app.route('/api/v1/pages/about-us/sections', methods=['GET'])
    def public_get_page_sections(slug=None):
        """Public API returning active sections and page metadata for a page ordered by sort_order."""
        target_slug = request.args.get('page') or slug or 'about-us'
        locale = request.args.get('locale') or request.args.get('lang') or 'en'
        sections = PageSection.all_for_page(page_slug=target_slug, include_inactive=False)
        formatted = [PageSection.to_localized_dict(s, locale=locale) for s in sections]

        page_obj = Page.find_by_slug(target_slug)
        page_data = page_obj.to_dict(locale=locale) if page_obj else {
            'slug': target_slug,
            'title': target_slug.replace('-', ' ').title(),
            'content': '',
            'meta_description': '',
            'seo_title': target_slug.replace('-', ' ').title()
        }

        return jsonify({
            'success': True,
            'page': page_data,
            'sections': formatted,
            'count': len(formatted)
        })

    # =========================================================================
    # 3. ENQUIRY / WHATSAPP BANNER SUBMISSION API (hdweb_enquiry table)
    # =========================================================================

    @app.route('/api/enquiry', methods=['POST'])
    @app.route('/api/v1/enquiry', methods=['POST'])
    def api_create_enquiry():
        """
        Receives quote & WhatsApp requests and saves all fields directly
        into the existing `hdweb_enquiry` table.
        """
        data = request.get_json(silent=True) or request.form.to_dict() or {}

        tyre_size = (data.get('tyre_size') or data.get('tyreSize') or '').strip()
        vehicle_raw = (data.get('vehicle') or data.get('carMake') or data.get('car_make') or '').strip()
        city = (data.get('city') or data.get('emirate') or '').strip()
        spec = (data.get('spec') or data.get('fitting') or '').strip()
        name = (data.get('name') or '').strip() or None
        email = (data.get('email') or '').strip() or None
        number = (data.get('number') or data.get('phone') or data.get('mobile') or '').strip() or None
        enquiry_for = (data.get('enquiry_for') or data.get('enquiryFor') or 'Tyre Quote (WhatsApp Home Banner)').strip()
        form_type = (data.get('form_type') or 'home_banner_whatsapp').strip()
        status = int(data.get('status', 0))

        # Check logged-in user in session if name, email, or mobile is not explicitly supplied
        if not name or not email or not number:
            name = name or session.get('name') or session.get('user_name') or session.get('customer_name')
            email = email or session.get('email') or session.get('user_email') or session.get('customer_email')
            number = number or session.get('phone') or session.get('mobile') or session.get('number') or session.get('customer_phone')

            sess_uid = session.get('user_id') or session.get('customer_id') or session.get('uid')
            if sess_uid:
                try:
                    conn_lookup = get_connection()
                    try:
                        with conn_lookup.cursor() as cur_lookup:
                            # 1. Check users table (e-commerce customer)
                            cur_lookup.execute("SELECT name, email, phone FROM users WHERE id = %s", (sess_uid,))
                            u_row = cur_lookup.fetchone()
                            if u_row:
                                name = name or u_row.get('name')
                                email = email or u_row.get('email')
                                number = number or u_row.get('phone')
                            else:
                                # 2. Check admin_users (administrator user)
                                cur_lookup.execute("SELECT name AS Name, email AS Email FROM admin_users WHERE id = %s", (sess_uid,))
                                u_tbl = cur_lookup.fetchone()
                                if u_tbl:
                                    name = name or u_tbl.get('Name')
                                    email = email or u_tbl.get('Email')
                    finally:
                        conn_lookup.close()
                except Exception as ex:
                    print("Session user lookup error in enquiry:", ex)

        # Extract make, model, year if available in vehicle string
        make = data.get('make')
        model = data.get('model')
        year = data.get('year')
        if vehicle_raw and (not make or not model):
            year_match = re.search(r'\b(19\d{2}|20\d{2})\b', vehicle_raw)
            if year_match:
                year = year or year_match.group(1)
            clean_no_year = re.sub(r'\b(19\d{2}|20\d{2})\b', '', vehicle_raw).strip()
            parts = [p for p in clean_no_year.split() if p]
            if parts and not make:
                make = parts[0].title()
            if len(parts) > 1 and not model:
                model = " ".join(parts[1:]).title()

        # Build message summary
        message = data.get('message')
        if not message:
            msg_parts = []
            if tyre_size:
                msg_parts.append(f"Tyre size: {tyre_size}")
            if vehicle_raw:
                msg_parts.append(f"Car: {vehicle_raw}")
            if city:
                msg_parts.append(f"Emirate: {city}")
            if spec:
                msg_parts.append(f"Fitting: {spec}")
            if name or email or number:
                user_info = []
                if name: user_info.append(f"Name: {name}")
                if email: user_info.append(f"Email: {email}")
                if number: user_info.append(f"Phone: {number}")
                msg_parts.append("User: " + ", ".join(user_info))
            message = "\n".join(msg_parts) if msg_parts else "WhatsApp Tyre Quote Request"

        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                sql = """
                    INSERT INTO hdweb_enquiry (
                        name, email, number, enquiry_for, message, status,
                        form_type, model, make, year, spec, current_insurance,
                        vehicle, tyre_size, city
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s
                    )
                """
                cursor.execute(sql, (
                    name,
                    email,
                    number,
                    enquiry_for,
                    message,
                    status,
                    form_type,
                    model,
                    make,
                    str(year) if year else None,
                    spec,
                    data.get('current_insurance'),
                    vehicle_raw or None,
                    tyre_size or None,
                    city or None
                ))
                conn.commit()
                new_id = cursor.lastrowid

            return jsonify({
                'success': True,
                'enquiry_id': new_id,
                'message': 'Enquiry successfully recorded in hdweb_enquiry.'
            }), 201
        except Exception as err:
            return jsonify({
                'success': False,
                'error': f"Failed to record enquiry: {str(err)}"
            }), 500
        finally:
            conn.close()


def register_api_routes(app):
    """Registers all API endpoints across tcsadmin, visionadmin, and the public client API."""
    register_tcsadmin_api_routes(app)
    register_visionadmin_api_routes(app)
    register_client_api_routes(app)
